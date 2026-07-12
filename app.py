import os
import sys
import re
import json
import time
import asyncio
import threading
import webbrowser
from datetime import datetime
from contextlib import asynccontextmanager
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_resource_path(relative_path: str) -> str:
    """Obtiene la ruta absoluta para un recurso, funciona en desarrollo y con PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_writeable_path(relative_path: str) -> str:
    """Obtiene la ruta absoluta para archivos persistentes de escritura en el directorio del ejecutable."""
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Directorios de datos
DATA_DIR = get_writeable_path("data")
EXTRACTIONS_DIR = os.path.join(DATA_DIR, "extractions")
CONSOLIDATED_DIR = os.path.join(DATA_DIR, "consolidated")
CONFIG_FILE = os.path.join(DATA_DIR, "config.json")
USERS_FILE = os.path.join(DATA_DIR, "users.json")
DICTIONARY_FILE = os.path.join(DATA_DIR, "dictionary.json")

def load_dictionary() -> Dict[str, Any]:
    if not os.path.exists(DICTIONARY_FILE):
        return {"categorias": {}, "marcas": {}}
    try:
        with open(DICTIONARY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        add_log("error", f"Error cargando dictionary.json: {str(e)}")
        return {"categorias": {}, "marcas": {}}

def get_provider_filepath(provider: Dict[str, Any]) -> str:
    """Obtiene la ruta absoluta para el archivo de salida de un proveedor."""
    provider_id = provider.get("id", "default")
    # Sanitizar provider_id para evitar inyección de rutas
    safe_id = os.path.basename(provider_id).replace("..", "").strip()
    if not safe_id or safe_id == ".":
        safe_id = "default"
    file_format = provider.get("file_format", "csv")
    if file_format not in ["csv", "xlsx"]:
        file_format = "csv"
    # Forzar a que el archivo esté estrictamente en EXTRACTIONS_DIR con un nombre seguro
    return os.path.join(EXTRACTIONS_DIR, f"{safe_id}.{file_format}")

import hashlib
import secrets

def hash_password(password: str) -> str:
    salt = os.urandom(16).hex()
    key = hashlib.pbkdf2_hmac(
        'sha256',
        password.encode('utf-8'),
        salt.encode('utf-8'),
        100000
    )
    return f"{salt}:{key.hex()}"

def verify_password(stored_password: str, provided_password: str) -> bool:
    try:
        salt, key_hex = stored_password.split(":")
        provided_key = hashlib.pbkdf2_hmac(
            'sha256',
            provided_password.encode('utf-8'),
            salt.encode('utf-8'),
            100000
        )
        return secrets.compare_digest(provided_key.hex(), key_hex)
    except Exception:
        return False

def load_users() -> Dict[str, str]:
    if not os.path.exists(USERS_FILE):
        return {}
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_users(users: Dict[str, str]):
    try:
        os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=2, ensure_ascii=False)
    except Exception as e:
        add_log("error", f"Error guardando users.json: {str(e)}")

def check_authentication(credentials: Optional[HTTPBasicCredentials] = Depends(HTTPBasic(auto_error=False))):
    correct_password = os.environ.get("ADMIN_PASSWORD")
    # Si no se define contraseña en el entorno, la autenticación está desactivada
    if not correct_password:
        return "developer"
        
    if not credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Autenticación requerida",
            headers={"WWW-Authenticate": "Basic"},
        )
        
    # 1. Verificar contra credenciales Root (entorno)
    correct_username = os.environ.get("ADMIN_USERNAME", "admin")
    is_correct_username = secrets.compare_digest(credentials.username, correct_username)
    is_correct_password = secrets.compare_digest(credentials.password, correct_password)
    
    if is_correct_username and is_correct_password:
        return credentials.username
        
    # 2. Verificar contra usuarios guardados en users.json
    users = load_users()
    if credentials.username in users:
        stored_hash = users[credentials.username]
        if verify_password(stored_hash, credentials.password):
            return credentials.username
            
    # Si ninguno coincide, levantar excepción
    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Credenciales incorrectas",
        headers={"WWW-Authenticate": "Basic"},
    )

def require_root(username: str = Depends(check_authentication)):
    correct_username = os.environ.get("ADMIN_USERNAME", "admin")
    # Si la contraseña no está definida en el entorno, permitimos el acceso a nivel de desarrollo
    if not os.environ.get("ADMIN_PASSWORD"):
        return username
    if username != correct_username:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acceso denegado: se requiere el rol de administrador principal (root)"
        )
    return username

# Estado global
is_monitoring = True
active_provider: Optional[Dict[str, Any]] = None
previous_clipboard = ""
last_sequence_number = 0
logs_buffer: List[Dict[str, Any]] = []
latest_log_index = 0
logs_lock = threading.Lock()

def add_log(msg_type: str, message: str, data: Any = None):
    global latest_log_index
    log_entry = {
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "type": msg_type, # 'info', 'success', 'warning', 'error'
        "message": message,
        "data": data
    }
    print(f"[{log_entry['timestamp']}] [{msg_type.upper()}] {message}", flush=True)
    with logs_lock:
        logs_buffer.append(log_entry)
        if len(logs_buffer) > 100:
            logs_buffer.pop(0)
        latest_log_index += 1

def load_config() -> Dict[str, Any]:
    global active_provider
    old_config = get_writeable_path("config.json")
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # Migrar config.json si existe en la ubicación antigua y no en la nueva
    if CONFIG_FILE != old_config and os.path.exists(old_config) and not os.path.exists(CONFIG_FILE):
        try:
            import shutil
            shutil.move(old_config, CONFIG_FILE)
            add_log("info", f"Migrado config.json de {old_config} a {CONFIG_FILE}")
        except Exception as e:
            add_log("warning", f"No se pudo migrar config.json automáticamente: {str(e)}")

    if not os.path.exists(CONFIG_FILE):
        config = {"active_provider_id": None, "providers": []}
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return config
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config = json.load(f)
            
        # Actualizar active_provider global
        providers = config.get("providers", [])
        active_id = config.get("active_provider_id")
        active_provider = next((p for p in providers if p["id"] == active_id), None)
        return config
    except Exception as e:
        add_log("error", f"Error cargando config.json: {str(e)}")
        raise RuntimeError(f"Error cargando config.json: {str(e)}") from e

def save_config(config: Dict[str, Any]):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except Exception as e:
        add_log("error", f"Error guardando config.json: {str(e)}")

def normalize_model_key(val: str) -> str:
    if not val:
        return ""
    # Convertir a minúsculas, quitar caracteres especiales y espacios
    return re.sub(r'[^a-zA-Z0-9]', '', str(val)).lower().strip()

def format_excel_file(filepath: str):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Fonts and Fills
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        body_font = Font(name="Segoe UI", size=10)
        cheap_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # Soft green
        cheap_font = Font(name="Segoe UI", size=10, bold=True, color="155724")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Headers
        headers = [cell.value for cell in ws[1]]
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Classify price columns by suffix
        price_cols_by_cat = {
            'general': [],
            'no_vat': [],
            'vat': [],
            'pvp': []
        }
        
        for idx, header in enumerate(headers):
            if header and header.startswith("Precio "):
                col_letter = get_column_letter(idx + 1)
                if header.endswith(" Sin IVA (€)"):
                    price_cols_by_cat['no_vat'].append((col_letter, idx + 1))
                elif header.endswith(" Con IVA (€)"):
                    price_cols_by_cat['vat'].append((col_letter, idx + 1))
                elif header.endswith(" PVP (€)"):
                    price_cols_by_cat['pvp'].append((col_letter, idx + 1))
                elif header.endswith(" (€)"):
                    price_cols_by_cat['general'].append((col_letter, idx + 1))
                    
        # Style rows
        num_rows = ws.max_row
        for row_idx in range(2, num_rows + 1):
            # Format body cells with font and border
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.border = thin_border
                
                # Align center for non-text columns
                if col_idx > 2:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
            # Highlight minimum price cell for each category
            for cat, cols in price_cols_by_cat.items():
                if len(cols) <= 1:
                    continue # No competition, no need to highlight
                    
                min_val = float('inf')
                min_cell = None
                
                for col_letter, col_idx in cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is not None:
                        try:
                            f_val = float(val)
                            if f_val > 0 and f_val < min_val:
                                min_val = f_val
                                min_cell = cell
                        except (ValueError, TypeError):
                            pass
                            
                if min_cell is not None and min_val != float('inf'):
                    min_cell.fill = cheap_fill
                    min_cell.font = cheap_font
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(filepath)
        wb.close()
        add_log("success", f"Archivo Excel formateado y coloreado correctamente: {filepath}")
    except Exception as e:
        add_log("error", f"Error formateando el archivo Excel {filepath}: {str(e)}")

def clean_price(price_str: str) -> float:
    if not price_str:
        return 0.0
    # Eliminar símbolos de moneda y espacios
    cleaned = re.sub(r'[^\d,.-]', '', price_str)
    # Estandarizar separador de decimales a punto
    if ',' in cleaned and '.' in cleaned:
        # Ambos presentes, por ejemplo: 1.250,50 o 1,250.50
        if cleaned.find('.') < cleaned.find(','):
            cleaned = cleaned.replace('.', '').replace(',', '.')
        else:
            cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        # Solo comas
        if cleaned.count(',') > 1:
            cleaned = cleaned.replace(',', '')
        else:
            parts = cleaned.split(',')
            if len(parts) == 2 and len(parts[1]) <= 2:
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
    elif '.' in cleaned:
        # Solo puntos
        if cleaned.count('.') > 1:
            cleaned = cleaned.replace('.', '')
        else:
            parts = cleaned.split('.')
            if len(parts) == 2 and len(parts[1]) == 3:
                # Ej: 1.250 -> miles (el electrodoméstico vale 1250, no 1.25)
                cleaned = cleaned.replace('.', '')
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def extract_products_adaptively(text: str, provider: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    results = []
    
    # 1. Encontrar todos los modelos posibles en el texto.
    # Exigimos que tengan al menos una letra y al menos un número, y mayúsculas/números/guiones (de 5 a 15 caracteres)
    # Ejemplo: 3TS3107BD, 3TS382B, WUU28T6XES, etc.
    model_pattern = r'\b(?=[A-Z0-9-]*[0-9])(?=[A-Z0-9-]*[A-Z])[A-Z0-9-]{5,15}\b'
    
    # Atributos comunes como capacidad (kg) y velocidad (rpm)
    attr_pattern = r'\b\d+\s*(?:kg|KG|Kg|rpm|RPM)\b'
    
    # Precios
    price_pattern = r'\b\d+(?:[\.,]\d+)*\s*€'
    
    # Extraer palabras clave del proveedor dinámicamente
    provider_keywords = set()
    if provider:
        # Palabras del nombre del proveedor
        prov_name = provider.get("name", "")
        for w in re.findall(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]{3,}', prov_name.lower()):
            provider_keywords.add(w)
            
        # Palabras de las etiquetas de producto/marca entrenadas
        sample_text = provider.get("sample_text", "")
        labels = provider.get("labels", [])
        product_label_texts = []
        for label in labels:
            name_lbl = label.get("name", "").lower()
            if any(term in name_lbl for term in ["product", "producto", "marca", "brand"]):
                start = label.get("start", 0)
                end = label.get("end", 0)
                if 0 <= start < end <= len(sample_text):
                    product_label_texts.append(sample_text[start:end].lower())
                    
        if not product_label_texts and sample_text:
            product_label_texts.append(sample_text.lower())
            
        for text_lbl in product_label_texts:
            for w in re.findall(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]{3,}', text_lbl):
                if w not in ["con", "del", "para", "por", "sus", "una", "uno", "los", "las", "les", "and", "the", "for"]:
                    provider_keywords.add(w)
                    
    if not provider_keywords:
        provider_keywords = {'lavadora', 'carga', 'frontal', 'balay', 'secadora', 'lavavajillas', 'electro'}
    
    # Procesar línea por línea
    lines = text.split('\n')
    spec_lines = []
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        if not line_stripped:
            continue
            
        # Descartamos líneas que terminan en puntos suspensivos ("...") o que son de ruido obvio
        if line_stripped.endswith('...') or re.search(r'\.\.\.\s*\d*$', line_stripped):
            continue
        if any(keyword in line_stripped.lower() for keyword in ['recíbelo entre', 'entrega garantizada', 'programas de lavado', 'tipo de instalación']):
            continue
            
        # Buscar modelos (sin re.IGNORECASE para evitar que coincidan palabras en minúsculas)
        models_in_line = re.findall(model_pattern, line_stripped)
        
        # Filtrar modelos falsos que son unidades de medida (ej: 1400RPM, 10KG)
        models_in_line = [m for m in models_in_line if not re.match(r'^\d+(?:KG|RPM|W|V|HZ|DB)$', m, re.IGNORECASE)]
        
        has_attrs = re.search(attr_pattern, line_stripped, re.IGNORECASE) is not None
        has_keywords = any(kw in line_stripped.lower() for kw in provider_keywords)
        
        # Si tiene modelo y atributos/palabras clave, la consideramos línea de especificación principal
        if models_in_line and (has_attrs or has_keywords):
            spec_lines.append({
                'index': i,
                'line': line_stripped,
                'model': models_in_line[0],
                'models': models_in_line
            })
            
    # Si no hay líneas de especificación con atributos, buscamos cualquier línea con un modelo válido
    if not spec_lines:
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            if line_stripped.endswith('...') or re.search(r'\.\.\.\s*\d*$', line_stripped):
                continue
            if any(keyword in line_stripped.lower() for keyword in ['recíbelo entre', 'entrega garantizada', 'programas de lavado', 'tipo de instalación']):
                continue
            models_in_line = re.findall(model_pattern, line_stripped)
            models_in_line = [m for m in models_in_line if not re.match(r'^\d+(?:KG|RPM|W|V|HZ|DB)$', m, re.IGNORECASE)]
            
            # Validación estricta para evitar falsos positivos de "letras y números sueltos"
            if models_in_line:
                line_has_keyword = any(kw in line_stripped.lower() for kw in provider_keywords)
                line_has_price = re.search(price_pattern, line_stripped) is not None
                if line_has_keyword or line_has_price:
                    spec_lines.append({
                        'index': i,
                        'line': line_stripped,
                        'model': models_in_line[0],
                        'models': models_in_line
                    })
                
    if not spec_lines:
        return []
        
    # 2. Para cada línea de especificación, delimitamos su bloque y extraemos los datos
    for idx, spec in enumerate(spec_lines):
        start_line_idx = spec['index']
        end_line_idx = spec_lines[idx+1]['index'] if idx + 1 < len(spec_lines) else len(lines)
        
        # El bloque de texto para este producto específico (hasta el inicio del siguiente producto)
        product_block = "\n".join(lines[start_line_idx:end_line_idx])
        
        spec_line = spec['line']
        model = spec['model']
        
        # Extraer Atributos Técnicos de la línea de especificación.
        attr_match = re.search(r'\b\d+\s*(?:kg|KG|Kg)\b', spec_line, re.IGNORECASE)
        if attr_match:
            attributes = spec_line[attr_match.start():].strip()
        else:
            attr_match_rpm = re.search(r'\b\d+\s*(?:rpm|RPM)\b', spec_line, re.IGNORECASE)
            if attr_match_rpm:
                attributes = spec_line[attr_match_rpm.start():].strip()
            else:
                model_pos = spec_line.find(model)
                if model_pos != -1:
                    attributes = spec_line[model_pos + len(model):].strip()
                else:
                    attributes = ""
                    
        # Limpiar atributos
        attributes = re.sub(r'\s+', ' ', attributes).strip()
        
        # Extraer Precio: primer precio en el bloque
        price_match = re.search(price_pattern, product_block)
        price = price_match.group(0) if price_match else "No disponible"
        
        # Limpiar precio de los atributos si se coló al final de la línea de especificación
        if price != "No disponible" and attributes:
            attributes_clean = attributes.replace(price, "").strip()
            # Eliminar guiones, comas o espacios finales sobrantes
            attributes = re.sub(r'\s+[-–,]\s*$', '', attributes_clean).strip()
            
        # Extraer Producto: es la línea de especificación menos el modelo y los atributos
        product = spec_line
        if attributes:
            # Quitamos los atributos originales de la línea
            product = product.replace(spec_line[attr_match.start():] if attr_match else attributes, '')
            
        product_parts = product.split(model)
        product_clean = " ".join([part.strip() for part in product_parts if part.strip()])
        
        product_clean = re.sub(r'\s+', ' ', product_clean).strip()
        product_clean = product_clean.strip(',').strip('-').strip()
        
        results.append({
            'product': product_clean,
            'model': model,
            'attributes': attributes,
            'price': price
        })
        
    return results

def find_model_column(df_columns, provider_fields) -> Optional[str]:
    # 1. Buscar nombres exactos comunes
    for col in df_columns:
        if col.lower() in ['model', 'modelo', 'sku', 'ref', 'referencia']:
            return col
    # 2. Buscar coincidencias parciales
    for col in df_columns:
        col_lower = col.lower()
        if 'model' in col_lower or 'sku' in col_lower or 'ref' in col_lower:
            return col
    # 3. Buscar en los campos configurados del proveedor
    for field in provider_fields:
        if field.lower() in ['model', 'modelo', 'sku', 'ref', 'referencia']:
            # Encontrar el nombre real de la columna en df
            for col in df_columns:
                if col.lower() == field.lower():
                    return col
    return None

def deduplicate_by_completeness(df: pd.DataFrame, model_col: str) -> pd.DataFrame:
    if df.empty or model_col not in df.columns:
        return df
        
    # Hacemos una copia del DataFrame para evitar mutar el original por referencia.
    df = df.copy()
    
    # Crear una copia temporal de la clave limpia para agrupar
    df['_temp_key'] = df[model_col].astype(str).str.strip().str.lower()
    
    # Excluir de la agrupación filas con clave vacía/nula
    df_valid = df[~df['_temp_key'].isin(['', 'nan', 'none'])].copy()
    df_invalid = df[df['_temp_key'].isin(['', 'nan', 'none'])].copy()
    
    if df_valid.empty:
        df = df.drop(columns=['_temp_key'], errors='ignore')
        return df
        
    # Calcular la "completitud" de cada fila: contar cuántas celdas no son nulas y no contienen marcadores de datos vacíos
    def count_valid_info(row):
        count = 0
        for col, val in row.items():
            if col in ['_temp_key', 'timestamp']:
                continue
            if pd.notnull(val):
                val_str = str(val).strip().lower()
                if val_str not in ["", "no disponible", "nan", "none", "null", "n/a", "-"]:
                    count += 1
        return count

    df_valid['_info_score'] = df_valid.apply(count_valid_info, axis=1)
    
    # Crear una columna con el índice original para mantener estabilidad
    df_valid['_orig_index'] = df_valid.index
    # Ordenar por '_info_score' descendente y '_orig_index' descendente (el más reciente primero en caso de empate)
    df_valid = df_valid.sort_values(by=['_info_score', '_orig_index'], ascending=[False, False])
    
    # Eliminar duplicados en la clave temporal, conservando la primera (la de más info y más reciente)
    df_clean = df_valid.drop_duplicates(subset=['_temp_key'], keep='first')
    
    # Restaurar el orden original por el índice
    df_clean = df_clean.sort_values(by='_orig_index')
    
    # Limpiar columnas temporales
    df_clean = df_clean.drop(columns=['_temp_key', '_info_score', '_orig_index'], errors='ignore')
    df_invalid = df_invalid.drop(columns=['_temp_key'], errors='ignore')
    
    return pd.concat([df_clean, df_invalid], ignore_index=True)

def process_text(text: str, provider: Dict[str, Any]):
    regex_pattern = provider.get("regex")
    if not regex_pattern:
        return
        
    try:
        # Filtrado previo por palabras clave requeridas del proveedor antes de procesar
        prov_name = provider.get("name", "")
        sample_text = provider.get("sample_text", "")
        
        required_keywords = set()
        for w in re.findall(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]{4,}', prov_name.lower()):
            required_keywords.add(w)
            
        labels = provider.get("labels", [])
        for label in labels:
            name_lbl = label.get("name", "").lower()
            if any(term in name_lbl for term in ["product", "producto", "marca", "brand"]):
                start = label.get("start", 0)
                end = label.get("end", 0)
                if 0 <= start < end <= len(sample_text):
                    for w in re.findall(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]{4,}', sample_text[start:end].lower()):
                        required_keywords.add(w)
                        
        if not required_keywords and sample_text:
            for w in re.findall(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]{4,}', sample_text.lower()):
                if w not in ["carga", "frontal", "para", "sobre", "este"]:
                    required_keywords.add(w)
                    
        if required_keywords:
            text_lower = text.lower()
            has_any_match = any(kw in text_lower for kw in required_keywords)
            
            # Permitir si contiene una estructura clara de producto (modelo + precio)
            has_product_structure = False
            model_match = re.search(r'\b(?=[A-Za-z0-9-]*\d)(?=[A-Za-z0-9-]*[A-Za-z])[A-Za-z0-9-]{5,15}\b', text)
            price_match = re.search(r'\b\d+(?:[\.,]\d+)*\s*€', text)
            if model_match and price_match:
                has_product_structure = True
                
            if not has_any_match and not has_product_structure:
                # Silenciosamente no hacemos nada, o informamos en logs sin saturar
                add_log("info", f"Texto ignorado: no contiene palabras clave de '{provider['name']}' ni estructura de producto.")
                return
                
        # Limpiar anclajes antiguos
        pattern = regex_pattern
        if pattern.startswith("^"):
            pattern = pattern[1:]
        if pattern.endswith("$"):
            pattern = pattern[:-1]
            
        matches = list(re.finditer(pattern, text))
        extracted_data_list = []
        
        # Limitar longitud máxima de coincidencia para evitar falsos positivos dispersos
        max_match_len = max(350, len(sample_text) * 3)
        valid_matches = []
        matched_intervals = []
        
        if matches:
            for match in matches:
                match_len = match.end() - match.start()
                if match_len <= max_match_len:
                    valid_matches.append(match.groupdict())
                    matched_intervals.append((match.start(), match.end()))
                else:
                    add_log("info", f"Coincidencia de expresión regular descartada por tamaño excesivo ({match_len} caracteres).")
                    
        # Obtener fragmentos de texto no coincidentes
        unmatched_segments = []
        last_idx = 0
        matched_intervals.sort(key=lambda x: x[0])
        for start, end in matched_intervals:
            if start > last_idx:
                segment = text[last_idx:start].strip()
                if segment:
                    unmatched_segments.append(segment)
            last_idx = end
        if last_idx < len(text):
            segment = text[last_idx:].strip()
            if segment:
                unmatched_segments.append(segment)

        # Si no hubo ninguna coincidencia válida con regex, el texto completo es un segmento no coincidente
        if not valid_matches and text.strip():
            unmatched_segments = [text.strip()]

        if valid_matches:
            extracted_data_list.extend(valid_matches)
            
        # Ejecutar extractor adaptativo para procesar todos los fragmentos que no coincidieron con la regex principal
        adaptive_extracted = []
        for segment in unmatched_segments:
            adaptive_matches = extract_products_adaptively(segment, provider)
            if adaptive_matches:
                expected_fields = provider.get("fields", [])
                for item in adaptive_matches:
                    data = {}
                    for field in expected_fields:
                        if field in ["product", "producto"]:
                            data[field] = item["product"]
                        elif field in ["model", "modelo"]:
                            data[field] = item["model"]
                        elif field in ["attributes", "atributos"]:
                            data[field] = item["attributes"]
                        elif field in ["price", "precio"]:
                            data[field] = item["price"]
                        elif field in ["price_no_vat", "precio_sin_iva"]:
                            data[field] = item["price"]
                        elif field in ["price_vat", "precio_con_iva"]:
                            data[field] = item["price"]
                        elif field in ["pvp", "precio_pvp"]:
                            data[field] = item["price"]
                        else:
                            data[field] = ""
                            
                    if not data:
                        data = item
                    adaptive_extracted.append(data)
                    
        if adaptive_extracted:
            if valid_matches:
                add_log("info", f"Se encontraron {len(valid_matches)} coincidencias exactas por plantilla y {len(adaptive_extracted)} adicionales usando extracción adaptativa inteligente.")
            else:
                add_log("info", "Iniciando extracción adaptativa inteligente en textos no coincidentes...")
            extracted_data_list.extend(adaptive_extracted)
        
        if extracted_data_list:
            added_count = 0
            for data in extracted_data_list:
                # Agregar timestamp
                data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Nombre de archivo de salida
                filepath = get_provider_filepath(provider)
                    
                file_format = provider.get("file_format", "csv")
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                
                df_new = pd.DataFrame([data])
                
                # Guardar en base al formato
                if file_format == "xlsx":
                    if os.path.exists(filepath):
                        try:
                            df_existing = pd.read_excel(filepath)
                            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                        except Exception:
                            df_combined = df_new
                    else:
                        df_combined = df_new
                        
                    # Deduplicación inteligente
                    model_col = find_model_column(df_combined.columns, provider.get("fields", []))
                    if model_col:
                        df_combined = deduplicate_by_completeness(df_combined, model_col)
                        
                    df_combined.to_excel(filepath, index=False)
                else:
                    # CSV
                    if os.path.exists(filepath):
                        try:
                            df_existing = pd.read_csv(filepath, encoding='utf-8-sig')
                            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                        except Exception:
                            df_combined = df_new
                    else:
                        df_combined = df_new
                        
                    # Deduplicación inteligente
                    model_col = find_model_column(df_combined.columns, provider.get("fields", []))
                    if model_col:
                        df_combined = deduplicate_by_completeness(df_combined, model_col)
                        
                    df_combined.to_csv(filepath, index=False, encoding='utf-8-sig')
                    
                product_name = data.get("product") or data.get("producto") or list(data.values())[0]
                add_log("success", f"Capturado y guardado: {product_name}", data)
                added_count += 1
                
            if added_count > 1:
                add_log("success", f"Se han procesado y guardado {added_count} productos del portapapeles.")
        else:
            # Ignorar de forma pasiva, pero registrar en el log informativo local
            add_log("info", f"Texto copiado ignorado (no coincide con la plantilla de '{provider['name']}')")
    except Exception as e:
        add_log("error", f"Error en el análisis del portapapeles: {str(e)}")

@asynccontextmanager
async def app_lifespan(app: FastAPI):
    global is_monitoring
    is_monitoring = True
    os.makedirs(EXTRACTIONS_DIR, exist_ok=True)
    os.makedirs(CONSOLIDATED_DIR, exist_ok=True)
    
    # Cargar configuración inicial
    load_config()
    
    yield
    
    is_monitoring = False

app = FastAPI(
    title="Garde Clipboard Parser",
    description="Local background app to parse clipboard data",
    lifespan=app_lifespan,
    dependencies=[Depends(check_authentication)]
)

# Servir estáticos (ruta compatible con PyInstaller)
static_dir = get_resource_path("static")
try:
    os.makedirs(static_dir, exist_ok=True)
except Exception:
    pass
app.mount("/static", StaticFiles(directory=static_dir), name="static")

# Modelos Pydantic
class LabelModel(BaseModel):
    name: str
    start: int
    end: int

class RegexGenerateRequest(BaseModel):
    raw_text: str
    labels: List[LabelModel]

class ProviderModel(BaseModel):
    id: str
    name: str
    regex: str
    fields: List[str]
    output_file: str
    file_format: str
    sample_text: str
    labels: List[LabelModel]

class SelectProviderRequest(BaseModel):
    provider_id: Optional[str] = None

class MergeRequest(BaseModel):
    files: List[str]
    merge_key: str
    output_filename: str

class ProcessTextRequest(BaseModel):
    text: str

class CreateUserRequest(BaseModel):
    username: str
    password: str

# Rutas API
@app.post("/api/process-text")
async def process_text_endpoint(req: ProcessTextRequest):
    global active_provider, is_monitoring
    if not is_monitoring:
        raise HTTPException(status_code=400, detail="El procesamiento online está desactivado.")
    if not active_provider:
        raise HTTPException(status_code=400, detail="Ningún proveedor activo seleccionado.")
    
    process_text(req.text, active_provider)
    return {"status": "success"}

@app.get("/api/users")
async def get_users(username: str = Depends(require_root)):
    users = load_users()
    # Devolver sólo los nombres de usuario, no sus contraseñas hash
    return {"users": list(users.keys())}

@app.post("/api/users")
async def create_user(req: CreateUserRequest, username: str = Depends(require_root)):
    cleaned_username = req.username.strip()
    if not cleaned_username:
        raise HTTPException(status_code=400, detail="El nombre de usuario no puede estar vacío.")
    if len(req.password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres.")
    
    # Comprobar que no coincida con el admin del entorno
    correct_username = os.environ.get("ADMIN_USERNAME", "admin")
    if cleaned_username.lower() == correct_username.lower():
        raise HTTPException(status_code=400, detail="No se puede crear un usuario con el nombre de administrador principal.")
        
    users = load_users()
    if cleaned_username in users:
        raise HTTPException(status_code=400, detail="El nombre de usuario ya está registrado.")
        
    hashed = hash_password(req.password)
    users[cleaned_username] = hashed
    save_users(users)
    add_log("info", f"Usuario estándar '{cleaned_username}' registrado por '{username}'.")
    return {"status": "success"}

@app.delete("/api/users/{username_to_delete}")
async def delete_user_route(username_to_delete: str, username: str = Depends(require_root)):
    users = load_users()
    if username_to_delete not in users:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    del users[username_to_delete]
    save_users(users)
    add_log("info", f"Usuario estándar '{username_to_delete}' eliminado por '{username}'.")
    return {"status": "success"}

@app.get("/")
async def get_index():
    index_path = os.path.join(get_resource_path("static"), "index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())
    return HTMLResponse("<h1>Dashboard no encontrado. Por favor, crea static/index.html</h1>")

@app.get("/api/status")
async def get_status(username: str = Depends(check_authentication)):
    global active_provider, is_monitoring
    is_root_user = False
    correct_username = os.environ.get("ADMIN_USERNAME", "admin")
    correct_password = os.environ.get("ADMIN_PASSWORD")
    
    # Si la autenticación está desactivada o el usuario es el admin/root
    if not correct_password or username == correct_username:
        is_root_user = True
        
    return {
        "is_monitoring": is_monitoring,
        "active_provider": active_provider,
        "has_active_provider": active_provider is not None,
        "username": username,
        "is_root": is_root_user,
        "auth_enabled": correct_password is not None
    }

@app.post("/api/status/toggle")
async def toggle_status():
    global is_monitoring, previous_clipboard, last_sequence_number
    is_monitoring = not is_monitoring
    state = "activado" if is_monitoring else "desactivado"
    add_log("info", f"Monitoreo de portapapeles {state} por el usuario.")
    if is_monitoring:
        # Al activar, reiniciamos el portapapeles previo para permitir re-capturar lo que ya esté copiado
        previous_clipboard = ""
        last_sequence_number = 0
    return {"is_monitoring": is_monitoring}

@app.get("/api/providers")
async def get_providers():
    config = load_config()
    return config

@app.post("/api/providers")
async def save_provider(provider: ProviderModel):
    global previous_clipboard, last_sequence_number
    config = load_config()
    providers = config.get("providers", [])
    
    # Buscar si ya existe y reemplazarlo, o añadirlo
    idx = next((i for i, p in enumerate(providers) if p["id"] == provider.id), -1)
    
    provider_dict = provider.model_dump()
    
    if idx >= 0:
        providers[idx] = provider_dict
        add_log("info", f"Plantilla de proveedor '{provider.name}' actualizada.")
    else:
        providers.append(provider_dict)
        add_log("info", f"Nueva plantilla de proveedor '{provider.name}' creada.")
        
    config["providers"] = providers
    save_config(config)
    load_config() # Recargar global
    previous_clipboard = ""
    last_sequence_number = 0
    return {"status": "success", "provider": provider_dict}

@app.delete("/api/providers/{provider_id}")
async def delete_provider(provider_id: str):
    config = load_config()
    providers = config.get("providers", [])
    
    filtered_providers = [p for p in providers if p["id"] != provider_id]
    if len(filtered_providers) == len(providers):
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        
    config["providers"] = filtered_providers
    if config.get("active_provider_id") == provider_id:
        config["active_provider_id"] = None
        
    save_config(config)
    load_config()
    add_log("info", f"Proveedor '{provider_id}' eliminado.")
    return {"status": "success"}

@app.post("/api/providers/select")
async def select_provider(req: SelectProviderRequest):
    global active_provider, previous_clipboard, last_sequence_number
    config = load_config()
    providers = config.get("providers", [])
    
    if req.provider_id is None:
        config["active_provider_id"] = None
        active_provider = None
        add_log("info", "Monitoreo desactivado: Ningún proveedor seleccionado.")
    else:
        # Verificar que existe
        prov = next((p for p in providers if p["id"] == req.provider_id), None)
        if not prov:
            raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        config["active_provider_id"] = req.provider_id
        active_provider = prov
        add_log("info", f"Proveedor seleccionado para captura: '{prov['name']}'.")
        
    save_config(config)
    # Al cambiar de proveedor, reiniciamos previous_clipboard para permitir capturar con las nuevas reglas
    previous_clipboard = ""
    last_sequence_number = 0
    return {"status": "success", "active_provider": active_provider}

@app.get("/api/providers/{provider_id}/data")
async def get_provider_data(provider_id: str):
    config = load_config()
    providers = config.get("providers", [])
    provider = next((p for p in providers if p["id"] == provider_id), None)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        
    filepath = get_provider_filepath(provider)
    if not os.path.exists(filepath):
        return {"columns": provider.get("fields", []) + ["timestamp"], "records": []}
        
    try:
        file_format = provider.get("file_format", "csv")
        if file_format == "xlsx":
            df = pd.read_excel(filepath)
        else:
            df = pd.read_csv(filepath, encoding='utf-8-sig')
            
        # Deduplicación al vuelo para limpiar cualquier duplicado residual en el archivo físico
        model_col = find_model_column(df.columns, provider.get("fields", []))
        if model_col:
            df_clean = deduplicate_by_completeness(df, model_col)
            if len(df_clean) < len(df):
                df = df_clean
                # Guardar el archivo de datos limpio de vuelta
                if file_format == "xlsx":
                    df.to_excel(filepath, index=False)
                else:
                    df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
        df = df.replace({pd.NA: None, float('nan'): None})
        
        # Construir registros con índice original de Pandas
        records = []
        for idx, row in df.iterrows():
            row_dict = row.to_dict()
            row_dict["_index"] = idx
            # Reemplazar NaN por None para evitar errores de serialización JSON
            row_dict = {k: (None if pd.isna(v) else v) for k, v in row_dict.items()}
            records.append(row_dict)
            
        columns = list(df.columns)
        return {"columns": columns, "records": records}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo archivo de datos: {str(e)}")

class DeleteRowRequest(BaseModel):
    index: int

@app.post("/api/providers/{provider_id}/clear")
async def clear_provider_data(provider_id: str):
    config = load_config()
    providers = config.get("providers", [])
    provider = next((p for p in providers if p["id"] == provider_id), None)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        
    filepath = get_provider_filepath(provider)
        
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            add_log("success", f"Se han eliminado todas las capturas del proveedor '{provider['name']}'.")
            return {"status": "success", "message": "Datos eliminados correctamente"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error al borrar el archivo de datos: {str(e)}")
    return {"status": "success", "message": "No había datos para eliminar"}

@app.post("/api/providers/{provider_id}/delete-row")
async def delete_provider_row(provider_id: str, req: DeleteRowRequest):
    config = load_config()
    providers = config.get("providers", [])
    provider = next((p for p in providers if p["id"] == provider_id), None)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        
    filepath = get_provider_filepath(provider)
        
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo de datos no encontrado")
        
    try:
        file_format = provider.get("file_format", "csv")
        if file_format == "xlsx":
            df = pd.read_excel(filepath)
        else:
            df = pd.read_csv(filepath, encoding='utf-8-sig')
            
        if req.index not in df.index:
            raise HTTPException(status_code=400, detail="Índice de fila no encontrado en el archivo")
            
        df = df.drop(index=req.index)
        
        if len(df) == 0:
            if os.path.exists(filepath):
                os.remove(filepath)
        else:
            if file_format == "xlsx":
                df.to_excel(filepath, index=False)
            else:
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                
        add_log("success", f"Captura eliminada correctamente.")
        return {"status": "success", "message": "Fila eliminada correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error eliminando la fila de datos: {str(e)}")

@app.post("/api/regex/generate")
async def generate_regex_pattern(req: RegexGenerateRequest):
    # Validar que tengamos al menos una etiqueta
    if not req.labels:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos una etiqueta para generar la expresión regular")
        
    raw_text = req.raw_text
    sorted_labels = sorted(req.labels, key=lambda x: x.start)
    
    parts = []
    
    # Función para determinar el patrón genérico de una etiqueta basada en su nombre o valor
    def get_generic_pattern(label_name: str, label_val: str, is_lazy: bool) -> str:
        name_lower = label_name.lower()
        val_clean = label_val.strip()
        
        # Si es un precio: contiene dígitos y opcionalmente símbolos de moneda
        if "precio" in name_lower or "price" in name_lower or "pvp" in name_lower or (re.search(r'\d', val_clean) and any(c in val_clean for c in ['€', '$', 'EUR', 'eur', 'Eur', 'usd', 'GBP', 'gbp'])):
            # Si el valor de muestra contiene un símbolo de moneda común, lo exigimos
            has_currency = any(c in val_clean for c in ['€', '$', 'EUR', 'eur', 'Eur', 'usd', 'GBP', 'gbp'])
            if has_currency:
                return r"\d+(?:[.,\d]*\d+)?\s*(?:€|EUR|eur|usd|\$|GBP|gbp)"
            else:
                return r"\d+(?:[.,\d]*\d+)?"
            
        # Si es un modelo/SKU: es alfanumérico y tiene cierta estructura
        if "modelo" in name_lower or "sku" in name_lower or (re.match(r'^[A-Za-z0-9-]+$', val_clean) and any(c.isdigit() for c in val_clean) and any(c.isalpha() for c in val_clean)):
            if any(c.isdigit() for c in val_clean) and any(c.isalpha() for c in val_clean):
                return r"\b(?=[A-Za-z0-9-]*\d)(?=[A-Za-z0-9-]*[A-Za-z])[A-Za-z0-9-]{4,25}\b"
            elif val_clean.isdigit():
                return r"\b\d{3,25}\b"
            else:
                return r"\b[A-Za-z0-9-]{3,25}\b"
            
        # Si es producto/marca o atributos:
        suffix = "?" if is_lazy else ""
        if "producto" in name_lower or "product" in name_lower or "marca" in name_lower or "brand" in name_lower:
            return f"[^\n]+{suffix}"
            
        if "atributo" in name_lower or "attr" in name_lower or "spec" in name_lower:
            return f"[^\n]+{suffix}"
            
        # Fallback por defecto:
        return f"[^\n]+{suffix}"
        
    def get_transition_pattern(literal: str) -> str:
        if not literal:
            return ""
        if not literal.strip():
            return r"\s+"
        parts = [re.escape(p) for p in literal.split()]
        return r"\s*" + r"\s+".join(parts) + r"\s*"
        
    # Construir la expresión regular iterando sobre las etiquetas y los textos intermedios
    last_idx = 0
    for i, label in enumerate(sorted_labels):
        start = label.start
        end = label.end
        name = label.name
        val = raw_text[start:end]
        
        literal_between = raw_text[last_idx:start]
        parts.append(get_transition_pattern(literal_between))
        
        # Determinar si este grupo es seguido por otro grupo en la misma línea
        is_lazy = False
        if i + 1 < len(sorted_labels):
            next_start = sorted_labels[i+1].start
            lit_after_this = raw_text[end:next_start]
            if "\n" not in lit_after_this:
                is_lazy = True
                
        # Añadir el grupo de captura genérico
        group_pattern = get_generic_pattern(name, val, is_lazy)
        parts.append(f"(?P<{name}>{group_pattern})")
        last_idx = end
        
    pattern = "".join(parts)
    
    # Validar usando re.search en el texto de muestra (sin anclajes strictly "^" y "$")
    try:
        match = re.search(pattern, raw_text)
        if match:
            extracted = match.groupdict()
            # Validar que los valores extraídos coincidan exactamente con lo que se seleccionó
            mismatches = []
            for label in sorted_labels:
                name = label.name
                expected_val = raw_text[label.start:label.end].strip()
                actual_val = extracted.get(name, "").strip()
                if expected_val != actual_val:
                    mismatches.append(f"Campo '{name}': esperado '{expected_val}', obtenido '{actual_val}'")
            
            if mismatches:
                return {
                    "status": "warning",
                    "regex": pattern,
                    "extracted": extracted,
                    "message": f"La expresión regular coincide pero los valores extraídos difieren: {'; '.join(mismatches)}"
                }
            else:
                return {
                    "status": "success",
                    "regex": pattern,
                    "extracted": extracted
                }
        else:
            return {
                "status": "warning",
                "regex": pattern,
                "message": "La expresión regular se generó pero no coincide con el texto de muestra. Por favor revisa los límites."
            }
    except Exception as e:
        return {
            "status": "error",
            "regex": pattern,
            "message": f"Error compilando la expresión regular: {str(e)}"
        }

@app.get("/api/extractions/files")
async def get_extraction_files():
    if not os.path.exists(EXTRACTIONS_DIR):
        return []
    files = []
    for f in os.listdir(EXTRACTIONS_DIR):
        if f.endswith('.csv') or f.endswith('.xlsx'):
            path = os.path.join(EXTRACTIONS_DIR, f)
            stat = os.stat(path)
            files.append({
                "filename": f,
                "size": stat.st_size,
                "last_modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    return files

@app.post("/api/extractions/merge")
async def merge_extractions(req: MergeRequest):
    if not req.files:
        raise HTTPException(status_code=400, detail="Debe seleccionar al menos un archivo")
        
    dfs = []
    
    for filename in req.files:
        filepath = os.path.join(EXTRACTIONS_DIR, filename)
        if not os.path.exists(filepath):
            continue
            
        # Leer archivo
        try:
            if filepath.endswith('.xlsx'):
                df = pd.read_excel(filepath)
            else:
                df = pd.read_csv(filepath, encoding='utf-8-sig')
        except Exception as e:
            add_log("warning", f"No se pudo leer el archivo '{filename}': {str(e)}")
            continue
            
        if df.empty:
            continue
            
        # Buscar columna de clave (Key)
        key_col = None
        for col in df.columns:
            if col.lower() == req.merge_key.lower():
                key_col = col
                break
        if not key_col:
            for col in df.columns:
                if req.merge_key.lower() in col.lower():
                    key_col = col
                    break
        if not key_col:
            key_col = df.columns[0] # Fallback a primera columna
            
        # Detectar columnas de precios (General, Sin IVA, Con IVA, PVP)
        general_price_col = None
        no_vat_price_col = None
        vat_price_col = None
        pvp_price_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'price' in col_lower or 'precio' in col_lower or 'pvp' in col_lower:
                if any(term in col_lower for term in ['sin_iva', 'sin iva', 'no_vat', 'no-vat']):
                    no_vat_price_col = col
                elif any(term in col_lower for term in ['con_iva', 'con iva', 'price_vat']) or col_lower.endswith('_vat'):
                    # Asegurar que no contenga 'no_vat'
                    if 'no_vat' not in col_lower and 'no-vat' not in col_lower:
                        vat_price_col = col
                elif 'pvp' in col_lower:
                    pvp_price_col = col
                else:
                    general_price_col = col
                    
        # Limpieza de clave con normalización difusa (remueve espacios, guiones y barras)
        df = df.dropna(subset=[key_col])
        df['_merge_key_clean'] = df[key_col].astype(str).apply(normalize_model_key)
        df = df[~df['_merge_key_clean'].isin(['', 'nan', 'none'])]
        
        # Limpieza de precios si existen
        rename_dict = {key_col: 'Producto / Modelo'}
        selected_cols = ['_merge_key_clean', 'Producto / Modelo']
        
        provider_name = filename.replace('.csv', '').replace('.xlsx', '').replace('_', ' ').title()
        
        if general_price_col:
            df['_price_clean'] = df[general_price_col].apply(lambda x: clean_price(str(x)) if pd.notnull(x) else 0.0)
            col_name = f'Precio {provider_name} (€)'
            rename_dict['_price_clean'] = col_name
            selected_cols.append(col_name)
            
        if no_vat_price_col:
            df['_price_no_vat_clean'] = df[no_vat_price_col].apply(lambda x: clean_price(str(x)) if pd.notnull(x) else 0.0)
            col_name_no_vat = f'Precio {provider_name} Sin IVA (€)'
            rename_dict['_price_no_vat_clean'] = col_name_no_vat
            selected_cols.append(col_name_no_vat)
            
        if vat_price_col:
            df['_price_vat_clean'] = df[vat_price_col].apply(lambda x: clean_price(str(x)) if pd.notnull(x) else 0.0)
            col_name_vat = f'Precio {provider_name} Con IVA (€)'
            rename_dict['_price_vat_clean'] = col_name_vat
            selected_cols.append(col_name_vat)
            
        if pvp_price_col:
            df['_price_pvp_clean'] = df[pvp_price_col].apply(lambda x: clean_price(str(x)) if pd.notnull(x) else 0.0)
            col_name_pvp = f'Precio {provider_name} PVP (€)'
            rename_dict['_price_pvp_clean'] = col_name_pvp
            selected_cols.append(col_name_pvp)
            
        # Desduplicar de forma inteligente manteniendo el registro con más información
        df_clean = deduplicate_by_completeness(df, key_col)
        df_clean = df_clean.rename(columns=rename_dict)
        
        dfs.append((provider_name, df_clean[selected_cols], {
            'has_general': general_price_col is not None,
            'has_no_vat': no_vat_price_col is not None,
            'has_vat': vat_price_col is not None,
            'has_pvp': pvp_price_col is not None
        }))
        
    if not dfs:
        raise HTTPException(status_code=400, detail="No se encontraron datos procesables en los archivos seleccionados")
        
    # Obtener todas las claves únicas
    keys_dict = {}
    for _, df, _ in dfs:
        for _, row in df.iterrows():
            keys_dict[row['_merge_key_clean']] = row['Producto / Modelo']
            
    merged_df = pd.DataFrame(list(keys_dict.items()), columns=['_merge_key_clean', 'Producto / Modelo'])
    
    price_cols = []
    price_no_vat_cols = []
    price_vat_cols = []
    price_pvp_cols = []
    
    for provider_name, df, col_flags in dfs:
        if col_flags['has_general']:
            col_name = f'Precio {provider_name} (€)'
            merged_df = pd.merge(merged_df, df[['_merge_key_clean', col_name]], on='_merge_key_clean', how='left')
            price_cols.append(col_name)
            
        if col_flags['has_no_vat']:
            col_name = f'Precio {provider_name} Sin IVA (€)'
            merged_df = pd.merge(merged_df, df[['_merge_key_clean', col_name]], on='_merge_key_clean', how='left')
            price_no_vat_cols.append(col_name)
            
        if col_flags['has_vat']:
            col_name = f'Precio {provider_name} Con IVA (€)'
            merged_df = pd.merge(merged_df, df[['_merge_key_clean', col_name]], on='_merge_key_clean', how='left')
            price_vat_cols.append(col_name)
            
        if col_flags.get('has_pvp'):
            col_name = f'Precio {provider_name} PVP (€)'
            merged_df = pd.merge(merged_df, df[['_merge_key_clean', col_name]], on='_merge_key_clean', how='left')
            price_pvp_cols.append(col_name)
            
    merged_df = merged_df.drop(columns=['_merge_key_clean'])
    
    # Calcular diferencia / oportunidad por categoría de precio
    def calculate_opportunity_for_cols(row, cols, label_suffix):
        prices = {}
        for col in cols:
            val = row[col]
            if pd.notnull(val) and val > 0:
                prices[col] = val
                
        if len(prices) == 0:
            return "Sin precios"
        if len(prices) == 1:
            prov = list(prices.keys())[0].replace('Precio ', '').replace(label_suffix, '').replace(' (€)', '').strip()
            return f"Solo en {prov}"
            
        sorted_prices = sorted(prices.items(), key=lambda x: x[1])
        cheapest_provider, cheapest_price = sorted_prices[0]
        second_provider, second_price = sorted_prices[1]
        
        cheapest_name = cheapest_provider.replace('Precio ', '').replace(label_suffix, '').replace(' (€)', '').strip()
        
        if cheapest_price == second_price:
            second_name = second_provider.replace('Precio ', '').replace(label_suffix, '').replace(' (€)', '').strip()
            return f"{cheapest_name} empata con {second_name}"
            
        diff_pct = ((second_price - cheapest_price) / second_price) * 100
        return f"{cheapest_name} ({diff_pct:.1f}% más barato)"
        
    if price_cols:
        merged_df['Diferencia / Oportunidad'] = merged_df.apply(
            lambda r: calculate_opportunity_for_cols(r, price_cols, ''), axis=1
        )
    if price_no_vat_cols:
        merged_df['Diferencia / Oportunidad Sin IVA'] = merged_df.apply(
            lambda r: calculate_opportunity_for_cols(r, price_no_vat_cols, 'Sin IVA'), axis=1
        )
    if price_vat_cols:
        merged_df['Diferencia / Oportunidad Con IVA'] = merged_df.apply(
            lambda r: calculate_opportunity_for_cols(r, price_vat_cols, 'Con IVA'), axis=1
        )
    if price_pvp_cols:
        merged_df['Diferencia / Oportunidad PVP'] = merged_df.apply(
            lambda r: calculate_opportunity_for_cols(r, price_pvp_cols, 'PVP'), axis=1
        )
        
    if not price_cols and not price_no_vat_cols and not price_vat_cols and not price_pvp_cols:
        merged_df['Diferencia / Oportunidad'] = "Sin precios cargados"
        
    # Guardar fusión consolidada
    out_filename = req.output_filename
    if not out_filename.endswith('.xlsx') and not out_filename.endswith('.csv'):
        out_filename += ".xlsx"
        
    out_filepath = os.path.join(CONSOLIDATED_DIR, out_filename)
    
    if out_filepath.endswith('.xlsx'):
        merged_df.to_excel(out_filepath, index=False)
        format_excel_file(out_filepath)
    else:
        merged_df.to_csv(out_filepath, index=False, encoding='utf-8-sig')
        
    # Reemplazar NaN por None para JSON serialización limpia
    merged_df = merged_df.replace({pd.NA: None, float('nan'): None})
    records = merged_df.to_dict(orient='records')
    
    add_log("success", f"Consolidación exitosa guardada en: {out_filename}")
    return {
        "status": "success",
        "file": out_filename,
        "columns": list(merged_df.columns),
        "data": records
    }

@app.post("/api/extractions/upload")
async def upload_extraction_files(files: List[UploadFile] = File(...)):
    os.makedirs(EXTRACTIONS_DIR, exist_ok=True)
    uploaded_files = []
    for file in files:
        filename = os.path.basename(file.filename)
        if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
            raise HTTPException(status_code=400, detail=f"Formato no válido: {filename}. Sólo se admiten archivos .csv y .xlsx")
            
        filepath = os.path.join(EXTRACTIONS_DIR, filename)
        try:
            with open(filepath, "wb") as f:
                content = await file.read()
                f.write(content)
            uploaded_files.append(filename)
            add_log("success", f"Archivo subido correctamente: {filename}")
        except Exception as e:
            add_log("error", f"Error al subir el archivo {filename}: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error escribiendo el archivo {filename}: {str(e)}")
            
    return {"status": "success", "uploaded": uploaded_files}

@app.delete("/api/extractions/files/{filename}")
async def delete_extraction_file(filename: str):
    filename = os.path.basename(filename)
    filepath = os.path.join(EXTRACTIONS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    try:
        os.remove(filepath)
        add_log("success", f"Archivo de extracción eliminado: {filename}")
        return {"status": "success", "message": f"Archivo {filename} eliminado"}
    except Exception as e:
        add_log("error", f"Error al eliminar el archivo {filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al eliminar el archivo: {str(e)}")

@app.get("/api/extractions/download/{provider_id}")
async def download_provider_extraction(provider_id: str):
    from fastapi.responses import FileResponse
    config = load_config()
    providers = config.get("providers", [])
    provider = next((p for p in providers if p["id"] == provider_id), None)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
        
    filepath = get_provider_filepath(provider)
        
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo de extracción no encontrado para este proveedor")
        
    filename = os.path.basename(filepath)
    return FileResponse(filepath, media_type="application/octet-stream", filename=filename)

@app.get("/api/consolidated/download/{filename}")
async def download_consolidated(filename: str):
    from fastapi.responses import FileResponse
    # Evitar Path Traversal sanitizando el nombre de archivo
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(CONSOLIDATED_DIR, safe_filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(filepath, media_type="application/octet-stream", filename=safe_filename)

@app.get("/api/stream")
async def sse_stream():
    async def event_generator():
        with logs_lock:
            current_logs = list(logs_buffer)
            last_sent_idx = latest_log_index
        # Envío inicial del búfer de logs
        yield f"data: {json.dumps({'type': 'init', 'logs': current_logs}, ensure_ascii=False)}\n\n"
        while True:
            await asyncio.sleep(0.5)
            with logs_lock:
                if last_sent_idx < latest_log_index:
                    diff = latest_log_index - last_sent_idx
                    start_pos = max(0, len(logs_buffer) - diff)
                    new_logs = list(logs_buffer)[start_pos:]
                    last_sent_idx = latest_log_index
                else:
                    new_logs = []
            
            if new_logs:
                for log in new_logs:
                    yield f"data: {json.dumps({'type': 'log', 'log': log}, ensure_ascii=False)}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'ping'}, ensure_ascii=False)}\n\n"
                
    return StreamingResponse(event_generator(), media_type="text/event-stream")

# --- MÓDULO DE MATRIZ DE STOCK E INVENTARIO ERP ---

EAN_CACHE_FILE = os.path.join(DATA_DIR, "ean_cache.json")
ean_cache_lock = threading.Lock()

def load_ean_cache() -> Dict[str, Any]:
    if not os.path.exists(EAN_CACHE_FILE):
        return {}
    try:
        with open(EAN_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_ean_cache(cache: Dict[str, Any]):
    try:
        with open(EAN_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception as e:
        add_log("error", f"Error guardando cache EAN: {str(e)}")

def query_ean_online(ean: str) -> Optional[Dict[str, Any]]:
    if not ean or ean == "N/D" or len(ean) < 8:
        return None
        
    with ean_cache_lock:
        cache = load_ean_cache()
        if ean in cache:
            return cache[ean]
        
    import urllib.request
    import urllib.error
    import json
    
    url = f"https://api.upcitemdb.com/prod/trial/lookup?upc={ean}"
    req = urllib.request.Request(
        url, 
        headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    )
    try:
        with urllib.request.urlopen(req, timeout=3) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data.get("code") == "OK" and data.get("items"):
                item = data["items"][0]
                result = {
                    "title": item.get("title") or "",
                    "brand": item.get("brand") or "",
                    "category": item.get("category") or ""
                }
                with ean_cache_lock:
                    cache = load_ean_cache()
                    cache[ean] = result
                    save_ean_cache(cache)
                return result
    except urllib.error.HTTPError as e:
        if e.code == 429:
            add_log("warning", f"Límite de API de EAN alcanzado (HTTP 429) al buscar {ean}")
        else:
            add_log("warning", f"Error HTTP {e.code} buscando EAN {ean}")
    except Exception as e:
        add_log("warning", f"Error de red/timeout buscando EAN {ean}: {str(e)}")
        
    return None

def check_category_prefix_rules(category_key: str, desc_lower: str) -> bool:
    clean_desc = re.sub(r'^[0-9\s\.,\-/()]+', '', desc_lower).strip()
    
    # Obtener la primera palabra (delimitada por espacios o puntuación/guiones)
    first_word_match = re.match(r'^([a-záéíóúñü]+)', clean_desc)
    if not first_word_match:
        return True
    first_word = first_word_match.group(1)
    
    # Mapeo de categorías con prefijos restrictivos obligatorios
    category_to_prefixes = {
        "Lavadoras-Secadoras": {"lavasecadora", "lavasecadoras", "lavadora"},
        "Lavadoras": {"lavadora", "lavadoras"},
        "Microondas": {"microondas", "micro"},
        "Hornos": {"horno", "hornos"},
        "Secadoras": {"secadora", "secadoras"},
        "Lavavajillas 45cm": {"lavavajillas", "lavaplatos"},
        "Lavavajillas 60cm": {"lavavajillas", "lavaplatos"},
        "Campanas": {"campana", "extractor", "extractora"},
        "Vitrocerámicas": {"placa", "vitro", "vitroceramica", "vitrocerámica", "induccion", "inducción", "encimera", "fuego", "fuegos"},
        "Calentadores": {"termo", "calentador", "calentadores"},
        "Ventiladores": {"ventilador", "ventiladores"},
        "Fregaderos": {"fregadero", "fregaderos"},
        "Climatizadores": {"climatizador", "aire"},
        "Frigoríficos": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"},
        "Frigo Combi": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"},
        "Frigo 2 puertas": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"},
        "Frigo 1 puerta": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"},
        "Frigos integrables": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"},
        "Frigos americanos": {"frigorifico", "frigorífico", "frigo", "frigos", "frigorificos", "frigoríficos", "congelador", "congeladores", "freezer", "freezers", "combi", "combis"}
    }
    
    if category_key in category_to_prefixes:
        return first_word in category_to_prefixes[category_key]
        
    return True

def classify_refrigerator(description: str) -> str:
    desc_lower = description.lower()
    
    # 1. Integrables
    if any(k in desc_lower for k in ["integrable", "integrables", "encastrable", "encastrables", "panelable", "panelables", "integrado", "integrados"]):
        return "Frigos integrables"
        
    # 2. Americanos
    if any(k in desc_lower for k in ["americano", "americanos", "side by side", "side-by-side", "multipuerta", "multi-puerta", "french door"]):
        return "Frigos americanos"
        
    # 3. 2 puertas
    if any(k in desc_lower for k in ["2 puertas", "dos puertas", "2 ptas", "2-puertas", "2ptas"]):
        return "Frigo 2 puertas"
        
    # 4. 1 puerta (incluye congeladores verticales y minibares)
    if any(k in desc_lower for k in ["1 puerta", "una puerta", "monopuerta", "1 pta", "1-puerta", "1pta", "cooler", "table top", "table-top", "congelador", "congeladores", "freezer", "freezers", "minibar", "mini bar", "mini-bar"]):
        return "Frigo 1 puerta"
        
    # 5. Combi
    if any(k in desc_lower for k in ["combi", "combis"]):
        return "Frigo Combi"
        
    return "Frigo Combi"


def extract_product_color(desc: str) -> str:
    """Extrae el color de un producto a partir de su descripción."""
    d = desc.lower()
    if any(k in d for k in ["inox", "inoxidable", "acero inoxidable", "stainless", "inox."]):
        return "Inox"
    if any(k in d for k in ["blanco", "white", "blanc"]):
        return "Blanco"
    if any(k in d for k in ["negro", "black", "noir"]):
        return "Negro"
    if any(k in d for k in ["titanio", "graphite", "grafito", "titanium"]):
        return "Titanio"
    return ""


def classify_lavavajillas(desc: str) -> str:
    """Clasifica un lavavajillas como 45cm o 60cm según su descripción."""
    d = desc.lower()
    # Patrones que indican 45cm
    if re.search(r'\b45\s*cm\b', d) or re.search(r'\b45\b', d):
        return "Lavavajillas 45cm"
    return "Lavavajillas 60cm"


def map_to_known_width(w_val: float) -> float:
    known_widths = [55.0, 59.5, 60.0, 70.0]
    best_w = known_widths[0]
    min_diff = float('inf')
    for kw in known_widths:
        diff = abs(w_val - kw)
        if diff < min_diff:
            min_diff = diff
            best_w = kw
    return best_w

def extract_frigo_medida(dim_text: str, is_americano: bool = False) -> str:
    # Buscar tres números de dimensiones separados por X o x (ej: 177,5X56X55)
    match = re.search(r'(?<![\d\.,])(\d+(?:[\.,]\d+)?)\s*[Xx]\s*(\d+(?:[\.,]\d+)?)\s*[Xx]\s*(\d+(?:[\.,]\d+)?)\b', dim_text)
    if not match:
        # Intentar también buscar dos números (por si acaso viene como 177,5x56)
        match_2 = re.search(r'(?<![\d\.,])(\d+(?:[\.,]\d+)?)\s*[Xx]\s*(\d+(?:[\.,]\d+)?)\b', dim_text)
        if match_2:
            g1 = match_2.group(1)
            g2 = match_2.group(2)
            
            # Reconstrucción de decimales si se cortó por un espacio antes del decimal (ej: '183, 5x59,5')
            start_idx = match_2.start(1)
            preceding_text = dim_text[:start_idx]
            prec_match = re.search(r'(\d+)[\.,]\s*$', preceding_text)
            if prec_match:
                g1 = f"{prec_match.group(1)}.{g1}"
                
            alto = float(g1.replace(',', '.'))
            ancho_val = float(g2.replace(',', '.'))
            
            if alto < 3.0:
                alto *= 100
            if alto >= 250.0:
                alto /= 10
            if ancho_val < 3.0:
                ancho_val *= 100
            if ancho_val >= 250.0:
                ancho_val /= 10
                
            if is_americano:
                ancho = ancho_val
            else:
                ancho = map_to_known_width(ancho_val)
            alto_str = str(alto).replace('.0', '').replace('.', ',')
            ancho_str = str(ancho).replace('.0', '').replace('.', ',')
            return f"{alto_str}x{ancho_str}"
        return "N/D"
        
    vals = [float(x.replace(',', '.')) for x in match.groups()]
    # Escalar si vienen en metros (ej. 1.86 -> 186)
    vals = [v * 100 if v < 3.0 else v for v in vals]
    # Escalar si vienen en milímetros (ej. 1860 -> 186, 682 -> 68.2)
    vals = [v / 10 if v >= 250.0 else v for v in vals]
    
    alto = max(vals)
    remaining = vals.copy()
    remaining.remove(alto)
    
    if is_americano:
        ancho = max(remaining)
    else:
        known_widths = [55.0, 59.5, 60.0, 70.0]
        best_val = remaining[0]
        min_diff = float('inf')
        best_mapped = known_widths[0]
        for r_val in remaining:
            for kw in known_widths:
                diff = abs(r_val - kw)
                if diff < min_diff:
                    min_diff = diff
                    best_val = r_val
                    best_mapped = kw
        ancho = best_mapped
        
    alto_str = str(alto).replace('.0', '').replace('.', ',')
    ancho_str = str(ancho).replace('.0', '').replace('.', ',')
    return f"{alto_str}x{ancho_str}"


def parse_erp_pdf(pdf_path: str) -> List[Dict[str, Any]]:
    import pypdf
    import re
    products = []
    
    def clean_numeric_token(token: str) -> Optional[float]:
        # Verificar que solo contiene dígitos, puntos, comas y opcionalmente signo
        if not re.match(r'^\d[\d\.,]*$', token):
            return None
        s = token
        if '.' in s and ',' in s:
            if s.find('.') < s.find(','): # Formato europeo: 11.908,59
                s = s.replace('.', '').replace(',', '.')
            else: # Formato americano: 11,908.59
                s = s.replace(',', '')
        elif ',' in s:
            if s.count(',') > 1:
                s = s.replace(',', '')
            else:
                s = s.replace(',', '.')
        elif '.' in s:
            if s.count('.') > 1:
                s = s.replace('.', '')
        try:
            return float(s)
        except ValueError:
            return None
    
    # Cargar diccionario semántico
    dct = load_dictionary()
    categorias_dict = dct.get("categorias", {})
    marcas_dict = dct.get("marcas", {})
    universal_units = ["kg", "l", "cubiertos", "servicios", "botellas", "zonas", "fuegos", "m3/h", "db", "w", "v", "rpm"]
    
    try:
        reader = pypdf.PdfReader(pdf_path)
        for page in reader.pages:
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            for line in lines:
                line_str = line.strip()
                if not line_str:
                    continue
                    
                # --- NUEVA LÓGICA PARA FORMATO TABULAR CON EURO (€) ---
                if '\u20ac' in line_str or '€' in line_str:
                    # Dividir la línea ignorando los caracteres de euro
                    clean_line = line_str.replace('\u20ac', '').replace('€', '')
                    tokens = clean_line.split()
                    if len(tokens) >= 2:
                        # 1. Identificar EAN/Código (token con >= 8 dígitos escaneando de derecha a izquierda)
                        code_ean = ""
                        ean_idx = -1
                        for idx in range(len(tokens) - 1, -1, -1):
                            tok = tokens[idx]
                            digits_only = re.sub(r'\D', '', tok)
                            if len(digits_only) >= 8:
                                code_ean = tok
                                ean_idx = idx
                                break
                                
                        # Extraer código y ean
                        if code_ean:
                            code_ean_digits = re.sub(r'\D', '', code_ean)
                            if len(code_ean_digits) >= 18:
                                code = code_ean_digits[:5]
                                ean = code_ean_digits[5:18]
                            elif len(code_ean_digits) >= 13:
                                ean = code_ean_digits[-13:]
                                code = code_ean_digits[:-13]
                            else:
                                ean = "N/D"
                                code = code_ean_digits
                        else:
                            ean = "N/D"
                            code = "N/D"
                            
                        # Crear lista de tokens sin el token EAN/código
                        remaining_tokens = tokens.copy()
                        if ean_idx != -1:
                            remaining_tokens.pop(ean_idx)
                            
                        if len(remaining_tokens) >= 1:
                            # 2. Identificar Coste y Stock
                            cost_val = None
                            stock_val = 1
                            desc_end_idx = len(remaining_tokens)
                            
                            last_token = remaining_tokens[-1]
                            last_num = clean_numeric_token(last_token)
                            
                            if last_num is not None:
                                if len(remaining_tokens) >= 2:
                                    prev_token = remaining_tokens[-2]
                                    prev_num = clean_numeric_token(prev_token)
                                    
                                    if prev_num is not None:
                                        if len(remaining_tokens) >= 3:
                                            third_token = remaining_tokens[-3]
                                            third_digits = re.sub(r'\D', '', third_token)
                                            if third_digits and len(third_digits) < 5:
                                                stock_val = int(third_digits)
                                                cost_val = prev_num
                                                desc_end_idx = -3
                                            else:
                                                cost_val = prev_num
                                                desc_end_idx = -2
                                        else:
                                            if last_num > prev_num:
                                                cost_val = last_num
                                                stock_val = int(prev_num) if prev_num > 0 else 1
                                            else:
                                                cost_val = prev_num
                                                stock_val = int(last_num) if last_num > 0 else 1
                                            desc_end_idx = -2
                                    else:
                                        cost_val = last_num
                                        stock_digits = re.sub(r'\D', '', prev_token)
                                        if stock_digits:
                                            stock_val = int(stock_digits)
                                            desc_end_idx = -2
                                        else:
                                            stock_val = 1
                                            desc_end_idx = -1
                                else:
                                    cost_val = last_num
                                    desc_end_idx = -1
                                    
                            if cost_val is not None:
                                # Extraer descripción (todos los tokens antes del bloque de stock/coste/ean)
                                desc_tokens = []
                                for idx, tok in enumerate(tokens):
                                    if idx == ean_idx:
                                        continue
                                    rem_idx = idx if (ean_idx == -1 or idx < ean_idx) else idx - 1
                                    if rem_idx >= len(remaining_tokens) + desc_end_idx:
                                        continue
                                    desc_tokens.append(tok)
                                    
                                raw_description = " ".join(desc_tokens).strip()
                                
                                # Buscar marca
                                brand = "Genérico"
                                desc_lower = raw_description.lower()
                                for brand_key, brand_variants in marcas_dict.items():
                                    if any(re.search(rf'\b{re.escape(v)}\b', desc_lower) for v in brand_variants):
                                        brand = brand_key
                                        break
                                        
                                # Buscar categoría — prioridad especial para Lavadoras-Secadoras
                                categoria = "Otros"
                                desc_lower_cat = raw_description.lower()
                                # Detectar lavasecadoras ANTES del bucle general
                                if ("lavadora" in desc_lower_cat and "secadora" in desc_lower_cat) or any(s in desc_lower_cat for s in ["lavasecadora", "lavasecadoras", "lavadora secadora", "washer dryer"]):
                                    categoria = "Lavadoras-Secadoras"
                                else:
                                    for cat_key, cat_val in categorias_dict.items():
                                        if cat_key in ("Lavadoras-Secadoras", "Lavavajillas 45cm", "Lavavajillas 60cm"):
                                            continue
                                        if not check_category_prefix_rules(cat_key, desc_lower):
                                            continue
                                        sinonimos = cat_val.get("sinonimos", [])
                                        if any(s in desc_lower for s in sinonimos):
                                            categoria = cat_key
                                            break
                                    # Subcategoría Lavavajillas por ancho
                                    if categoria == "Lavavajillas":
                                        categoria = classify_lavavajillas(raw_description)
                                        
                                # Unidades para capacidad
                                cat_info = categorias_dict.get(categoria, {})
                                attrs_info = cat_info.get("atributos_clave", {})
                                size_units = []
                                for attr_name, units in attrs_info.items():
                                    if attr_name in ["capacidad", "servicios", "extraccion", "zonas"]:
                                        size_units.extend(units)
                                if not size_units:
                                    size_units = universal_units
                                    
                                # Intentar extraer dimensiones físicas si es categoría de frigoríficos
                                is_frigo_category = "frigo" in categoria.lower() or "frigorífico" in categoria.lower() or "frigorico" in categoria.lower()
                                capacidad = "N/D"
                                if is_frigo_category:
                                    is_americano = "americano" in categoria.lower()
                                    capacidad = extract_frigo_medida(raw_description, is_americano)
                                    
                                desc_words = []
                                
                                i = 0
                                while i < len(desc_tokens):
                                    token = desc_tokens[i]
                                    token_upper = token.upper()
                                    is_tech_spec = False
                                    
                                    # Caso 1: Unidad pegada
                                    unit_match = re.match(r'^(\d+[\.,]?\d*)([A-Z][A-Z0-9/]*)$', token_upper)
                                    if unit_match:
                                        val = unit_match.group(1)
                                        unit = unit_match.group(2).lower()
                                        if unit in size_units:
                                            unit_label = "kg" if unit in ["kg", "kilogramos", "kilos"] else ("L" if unit in ["l", "litros", "lts"] else unit)
                                            if not (is_frigo_category and capacidad != "N/D"):
                                                capacidad = f"{val} {unit_label}"
                                            is_tech_spec = True
                                        elif any(unit in unit_list for unit_list in attrs_info.values()) or unit in ["w", "v", "hz", "db", "rpm"]:
                                                is_tech_spec = True
                                                
                                    # Caso 2: Unidad separada
                                    elif i + 1 < len(desc_tokens):
                                        next_token_lower = desc_tokens[i+1].lower()
                                        if re.match(r'^\d+[\.,]?\d*$', token):
                                            if next_token_lower in size_units:
                                                unit_label = "kg" if next_token_lower in ["kg", "kilogramos", "kilos"] else ("L" if next_token_lower in ["l", "litros", "lts"] else next_token_lower)
                                                if not (is_frigo_category and capacidad != "N/D"):
                                                    capacidad = f"{token} {unit_label}"
                                                is_tech_spec = True
                                                i += 1
                                            elif any(next_token_lower in unit_list for unit_list in attrs_info.values()) or next_token_lower in ["w", "v", "hz", "db", "rpm"]:
                                                is_tech_spec = True
                                                i += 1
                                                
                                    if is_tech_spec:
                                        i += 1
                                        continue
                                        
                                    if token.lower() not in ["de", "con", "el", "la", "en", "para"]:
                                        desc_words.append(token)
                                    i += 1
                                    
                                # Fallback adaptativo para marca/categoría si son Genéricos/Otros
                                if brand == "Genérico" and categoria == "Otros":
                                    clean_words = []
                                    for word in desc_words:
                                        word_clean = word.strip().strip(",.-/()").title()
                                        if len(word_clean) >= 3:
                                            clean_words.append(word_clean)
                                    if len(clean_words) >= 2:
                                        categoria = clean_words[0]
                                        brand = clean_words[1]
                                    elif len(clean_words) == 1:
                                        categoria = clean_words[0]
                                        brand = "Genérico"
                                else:
                                    if brand == "Genérico":
                                        for word in desc_words:
                                            word_clean = word.strip().strip(",.-/()").title()
                                            if len(word_clean) >= 3:
                                                is_cat_synonym = False
                                                for cat_val in categorias_dict.values():
                                                    if word_clean.lower() in cat_val.get("sinonimos", []):
                                                        is_cat_synonym = True
                                                        break
                                                if not is_cat_synonym:
                                                    brand = word_clean
                                                    break
                                                    
                                    if categoria == "Otros":
                                        for word in desc_words:
                                            word_clean = word.strip().strip(",.-/()").title()
                                            if len(word_clean) >= 3 and word_clean.lower() != brand.lower():
                                                categoria = word_clean
                                                break
                                                
                                desc = " ".join(desc_words)
                                desc = re.sub(r'\s+', ' ', desc).strip()
                                if not desc:
                                    desc = f"Electrodoméstico {brand}"
                                    
                                # Buscar SKU/Modelo en la descripción limpia o en los tokens de la descripción
                                model_match = re.search(r'\b(?=[A-Z0-9-]*[0-9])(?=[A-Z0-9-]*[A-Z])[A-Z0-9-]{4,15}\b', raw_description.upper())
                                if model_match:
                                    model = model_match.group(0)
                                else:
                                    model = code

                                # Asegurar que los frigoríficos se subdividen y clasifican por medida
                                is_frigo = (
                                    categoria == "Frigoríficos" or 
                                    "frigo" in categoria.lower() or 
                                    "frigorífico" in categoria.lower() or 
                                    "frigorico" in categoria.lower() or
                                    any(w in desc.lower() for w in ["congelador", "freezer", "combi"])
                                )
                                if is_frigo:
                                    is_americano = "americano" in categoria.lower() or any(w in desc.lower() for w in ["americano", "americanos", "side by side", "multipuerta"])
                                    medida = extract_frigo_medida(desc, is_americano)
                                    if medida == "N/D":
                                        medida = extract_frigo_medida(raw_description, is_americano)
                                    
                                    categoria = classify_refrigerator(desc)
                                    if medida != "N/D":
                                        capacidad = medida
                                    
                                color = extract_product_color(raw_description)
                                products.append({
                                    "sku": model,
                                    "ean": ean,
                                    "code": code,
                                    "brand": brand,
                                    "category": categoria,
                                    "description": desc,
                                    "capacity": capacidad,
                                    "color": color,
                                    "stock": stock_val,
                                    "cost": cost_val
                                })
                        continue
                        
                # Saltar líneas de encabezados o metadatos de página
                if any(t in line_str.lower() for t in ["total general", "subtotal", "valoracion", "valoración", "pagina", "página", "listado de stock"]):
                    continue
                
                # 1. Buscar Modelo/SKU de electrodoméstico
                model_match = re.search(r'\b(?=[A-Z0-9-]*[0-9])(?=[A-Z0-9-]*[A-Z])[A-Z0-9-]{5,15}\b', line_str.upper())
                if not model_match:
                    continue
                    
                model = model_match.group(0)
                
                # Clasificación de categoría basada en el diccionario — prioridad para Lavadoras-Secadoras
                desc_lower = line_str.lower()
                categoria = "Otros"
                if ("lavadora" in desc_lower and "secadora" in desc_lower) or any(s in desc_lower for s in ["lavasecadora", "lavasecadoras", "lavadora secadora", "washer dryer"]):
                    categoria = "Lavadoras-Secadoras"
                else:
                    for cat_key, cat_val in categorias_dict.items():
                        if cat_key in ("Lavadoras-Secadoras", "Lavavajillas 45cm", "Lavavajillas 60cm"):
                            continue
                        if not check_category_prefix_rules(cat_key, desc_lower):
                            continue
                        sinonimos = cat_val.get("sinonimos", [])
                        if any(s in desc_lower for s in sinonimos):
                            categoria = cat_key
                            break
                    # Subcategoría Lavavajillas por ancho
                    if categoria == "Lavavajillas":
                        categoria = classify_lavavajillas(line_str)
                
                # Dividir la línea en tokens de texto limpios
                tokens = line_str.split()
                
                # Variables para extraer
                capacidad = "N/D"
                
                # Intentar extraer dimensiones físicas si es categoría de frigoríficos
                is_frigo_category = "frigo" in categoria.lower() or "frigorífico" in categoria.lower() or "frigorico" in categoria.lower()
                if is_frigo_category:
                    is_americano = "americano" in categoria.lower()
                    capacidad = extract_frigo_medida(line_str, is_americano)
                
                brand = "Genérico"
                candidate_numbers = []
                desc_words = []
                
                # Detectar marca usando marcas_dict (con límites de palabra para evitar colisiones)
                for brand_key, brand_variants in marcas_dict.items():
                    if any(re.search(rf'\b{re.escape(v)}\b', desc_lower) for v in brand_variants):
                        brand = brand_key
                        break
                
                # Obtener unidades de capacidad/medida asociadas a esta categoría
                cat_info = categorias_dict.get(categoria, {})
                attrs_info = cat_info.get("atributos_clave", {})
                size_units = []
                for attr_name, units in attrs_info.items():
                    if attr_name in ["capacidad", "servicios", "extraccion", "zonas"]:
                        size_units.extend(units)
                
                # Si no hay unidades específicas en el diccionario, usar el listado universal como fallback
                if not size_units:
                    size_units = universal_units
                
                # Procesar tokens de izquierda a derecha
                i = 0
                while i < len(tokens):
                    token = tokens[i]
                    token_upper = token.upper()
                    
                    # Si es el SKU, lo saltamos
                    if token_upper == model:
                        i += 1
                        continue
                        
                    is_tech_spec = False
                    
                    # Caso 1: Unidad pegada (ej: "8KG", "1200RPM", "368L")
                    unit_match = re.match(r'^(\d+[\.,]?\d*)([A-Z][A-Z0-9/]*)$', token_upper)
                    if unit_match:
                        val = unit_match.group(1)
                        unit = unit_match.group(2).lower()
                        
                        if unit in size_units:
                            unit_label = "kg" if unit in ["kg", "kilogramos", "kilos"] else ("L" if unit in ["l", "litros", "lts"] else unit)
                            if not (is_frigo_category and capacidad != "N/D"):
                                capacidad = f"{val} {unit_label}"
                            is_tech_spec = True
                        elif any(unit in unit_list for unit_list in attrs_info.values()) or unit in ["w", "v", "hz", "db", "rpm"]:
                            is_tech_spec = True
                        
                    # Caso 2: Unidad separada (ej: i es número, i+1 es unidad)
                    elif i + 1 < len(tokens):
                        next_token_lower = tokens[i+1].lower()
                        if re.match(r'^\d+[\.,]?\d*$', token):
                            if next_token_lower in size_units:
                                unit_label = "kg" if next_token_lower in ["kg", "kilogramos", "kilos"] else ("L" if next_token_lower in ["l", "litros", "lts"] else next_token_lower)
                                if not (is_frigo_category and capacidad != "N/D"):
                                    capacidad = f"{token} {unit_label}"
                                is_tech_spec = True
                                i += 1 # Consumir unidad
                            elif any(next_token_lower in unit_list for unit_list in attrs_info.values()) or next_token_lower in ["w", "v", "hz", "db", "rpm"]:
                                is_tech_spec = True
                                i += 1
                                
                    if is_tech_spec:
                        i += 1
                        continue
                        
                    # Si no es atributo técnico, comprobar si es un número puro candidato a stock/coste
                    number_match = re.match(r'^\d+(?:[\.,]\d+)?$', token)
                    if number_match:
                        try:
                            val = float(token.replace(",", "."))
                            candidate_numbers.append((token, val))
                        except ValueError:
                            pass
                    else:
                        # Si es palabra de descripción
                        if token.lower() not in ["de", "con", "el", "la", "en", "para"]:
                            desc_words.append(token)
                            
                    i += 1
                
                # Determinar stock y costo basados en los números candidatos encontrados
                # Filtrar candidatos: eliminar códigos/EANs (ej: con 8+ dígitos o >= 100,000)
                clean_candidates = []
                ean_val = "N/D"
                for c_tok, c_val in candidate_numbers:
                    digits_only = re.sub(r'\D', '', c_tok)
                    if len(digits_only) >= 8 or c_val >= 100000:
                        if ean_val == "N/D" and len(digits_only) >= 8:
                            ean_val = digits_only
                        continue
                    clean_candidates.append((c_tok, c_val))
                    
                # Determinar stock y costo basados en los candidatos limpios
                stock = 0
                cost = 0.0
                
                if len(clean_candidates) == 1:
                    token, val = clean_candidates[0]
                    if val > 100:
                        cost = val
                        stock = 1
                    else:
                        stock = int(val)
                elif len(clean_candidates) == 2:
                    c1_token, val1 = clean_candidates[0]
                    c2_token, val2 = clean_candidates[1]
                    
                    has_decimal1 = ("," in c1_token) or ("." in c1_token)
                    has_decimal2 = ("," in c2_token) or ("." in c2_token)
                    
                    if has_decimal1 and not has_decimal2:
                        cost = val1
                        stock = int(val2)
                    elif not has_decimal1 and has_decimal2:
                        stock = int(val1)
                        cost = val2
                    elif has_decimal1 and has_decimal2:
                        # Ambos tienen decimales (ej: Coste y Total). Calcular stock = Total / Coste
                        if val1 > 0 and val2 > val1:
                            cost = val1
                            stock = int(round(val2 / val1))
                        elif val2 > 0 and val1 > val2:
                            cost = val2
                            stock = int(round(val1 / val2))
                        else:
                            cost = val1
                            stock = 1
                    else:
                        # Ninguno tiene decimales
                        if val1 > val2:
                            if val1 > 100:
                                cost = val1
                                stock = int(val2)
                            else:
                                stock = int(val1)
                                cost = val2
                        else:
                            if val2 > 100:
                                cost = val2
                                stock = int(val1)
                            else:
                                stock = int(val2)
                                cost = val1
                elif len(clean_candidates) >= 3:
                    # Intentar buscar trío Stock, Coste, Total
                    c1_token, c1_val = clean_candidates[-3]
                    c2_token, c2_val = clean_candidates[-2]
                    c3_token, c3_val = clean_candidates[-1]
                    
                    if c1_val > 0 and c2_val > 0 and abs(c1_val * c2_val - c3_val) < max(5.0, c3_val * 0.02):
                        stock = int(c1_val)
                        cost = c2_val
                    elif c1_val > 0 and c2_val > 0 and abs(c2_val * c1_val - c3_val) < max(5.0, c3_val * 0.02):
                        stock = int(c2_val)
                        cost = c1_val
                    else:
                        # Fallback a los últimos 2 usando la lógica de 2 candidatos
                        val1 = c2_val
                        val2 = c3_val
                        c1_token = c2_token
                        c2_token = c3_token
                        has_decimal1 = ("," in c1_token) or ("." in c1_token)
                        has_decimal2 = ("," in c2_token) or ("." in c2_token)
                        if has_decimal1 and not has_decimal2:
                            cost = val1
                            stock = int(val2)
                        elif not has_decimal1 and has_decimal2:
                            stock = int(val1)
                            cost = val2
                        elif has_decimal1 and has_decimal2:
                            if val1 > 0 and val2 > val1:
                                cost = val1
                                stock = int(round(val2 / val1))
                            elif val2 > 0 and val1 > val2:
                                cost = val2
                                stock = int(round(val1 / val2))
                            else:
                                cost = val1
                                stock = 1
                        else:
                            if val1 > val2:
                                cost = val1
                                stock = int(val2)
                            else:
                                cost = val2
                                stock = int(val1)
                                
                desc = " ".join(desc_words)
                desc = re.sub(r'\s+', ' ', desc).strip()
                
                # Fallback adaptativo conjunto para marca y categoría si son genéricos/Otros
                if brand == "Genérico" and categoria == "Otros":
                    clean_words = []
                    for word in desc_words:
                        word_clean = word.strip().strip(",.-/()").title()
                        if len(word_clean) >= 3:
                            clean_words.append(word_clean)
                    if len(clean_words) >= 2:
                        categoria = clean_words[0]
                        brand = clean_words[1]
                    elif len(clean_words) == 1:
                        categoria = clean_words[0]
                        brand = "Genérico"
                else:
                    # Fallback individual para marca si sigue siendo genérica
                    if brand == "Genérico":
                        for word in desc_words:
                            word_clean = word.strip().strip(",.-/()").title()
                            if len(word_clean) >= 3:
                                is_cat_synonym = False
                                for cat_val in categorias_dict.values():
                                    if word_clean.lower() in cat_val.get("sinonimos", []):
                                        is_cat_synonym = True
                                        break
                                if not is_cat_synonym:
                                    brand = word_clean
                                    break
                                    
                    # Fallback adaptativo para categoría si sigue siendo Otros
                    if categoria == "Otros":
                        for word in desc_words:
                            word_clean = word.strip().strip(",.-/()").title()
                            if len(word_clean) >= 3 and word_clean.lower() != brand.lower():
                                categoria = word_clean
                                break

                if not desc:
                    desc = f"Electrodoméstico {brand}"
                    
                # Asegurar que los frigoríficos se subdividen y clasifican por medida
                is_frigo = (
                    categoria == "Frigoríficos" or 
                    "frigo" in categoria.lower() or 
                    "frigorífico" in categoria.lower() or 
                    "frigorico" in categoria.lower() or
                    any(w in desc.lower() for w in ["congelador", "freezer", "combi"])
                )
                if is_frigo:
                    is_americano = "americano" in categoria.lower() or any(w in desc.lower() for w in ["americano", "americanos", "side by side", "multipuerta"])
                    medida = extract_frigo_medida(desc, is_americano)
                    if medida == "N/D":
                        medida = extract_frigo_medida(line_str, is_americano)
                    
                    categoria = classify_refrigerator(desc)
                    if medida != "N/D":
                        capacidad = medida

                color = extract_product_color(line_str)
                products.append({
                    "sku": model,
                    "ean": ean_val,
                    "brand": brand,
                    "category": categoria,
                    "description": desc,
                    "capacity": capacidad,
                    "color": color,
                    "stock": stock,
                    "cost": cost
                })
    except Exception as e:
        add_log("error", f"Error parseando PDF: {str(e)}")
        
    # Enriquecer productos usando los EANs online de forma asíncrona/concurrente
    valid_products = [p for p in products if p.get("ean") and p["ean"] != "N/D" and len(p["ean"]) >= 8]
    if valid_products:
        from concurrent.futures import ThreadPoolExecutor
        cache = load_ean_cache()
        to_lookup = [p["ean"] for p in valid_products if p["ean"] not in cache]
        
        # Limitar número de peticiones por subida para no saturar la API
        max_lookups = 40
        to_lookup = to_lookup[:max_lookups]
        
        if to_lookup:
            add_log("info", f"Buscando {len(to_lookup)} EANs nuevos online (límite: {max_lookups})...")
            try:
                with ThreadPoolExecutor(max_workers=5) as executor:
                    executor.map(query_ean_online, to_lookup)
            except Exception as ex:
                add_log("warning", f"Error en ejecución concurrente de EAN: {str(ex)}")
                
        # Aplicar resultados de la caché
        cache = load_ean_cache()
        for p in valid_products:
            ean = p["ean"]
            if ean in cache:
                online_info = cache[ean]
                online_brand = online_info.get("brand")
                online_title = online_info.get("title")
                online_cat = online_info.get("category")
                
                # Enriquecer Marca si es genérica
                if (p.get("brand") == "Genérico" or not p.get("brand")) and online_brand:
                    matched_brand = None
                    for b_key, b_variants in marcas_dict.items():
                        if any(v == online_brand.lower() for v in b_variants):
                            matched_brand = b_key
                            break
                    if matched_brand:
                        p["brand"] = matched_brand
                    else:
                        p["brand"] = online_brand.strip().title()
                        
                # Enriquecer Categoría si es Otros
                if p.get("category") == "Otros" or not p.get("category"):
                    search_text = f"{online_title} {online_cat}".lower()
                    for cat_key, cat_val in categorias_dict.items():
                        sinonimos = cat_val.get("sinonimos", [])
                        if any(s in search_text for s in sinonimos):
                            p["category"] = cat_key
                            break
                            
                # Si ahora es frigorífico, asegurar subdivisión y medida
                cat_current = p.get("category", "")
                desc_current = p.get("description", "")
                is_frigo = (
                    cat_current == "Frigoríficos" or 
                    "frigo" in cat_current.lower() or 
                    "frigorífico" in cat_current.lower() or 
                    "frigorico" in cat_current.lower() or
                    any(w in desc_current.lower() for w in ["congelador", "freezer", "combi"])
                )
                if is_frigo:
                    is_americano = "americano" in cat_current.lower() or any(w in desc_current.lower() for w in ["americano", "americanos", "side by side", "multipuerta"])
                    medida = extract_frigo_medida(desc_current, is_americano)
                    p["category"] = classify_refrigerator(desc_current)
                    if medida != "N/D":
                        p["capacity"] = medida
                            
    return products

def get_stock_matrix_data(category: str = "Lavadoras", color_filter: str = "") -> Dict[str, Any]:
    inventory_file = os.path.join(DATA_DIR, "stock", "inventory.json")
    if not os.path.exists(inventory_file):
        return {
            "categories": [],
            "selected_category": category,
            "brands": [],
            "brands_dist": [],
            "capacities": [],
            "cells": [],
            "kpis": {"total_value": 0.0, "total_references": 0, "total_stock": 0, "coverage_pct": 0.0},
            "alerts": []
        }
        
    try:
        with open(inventory_file, "r", encoding="utf-8") as f:
            all_products = json.load(f)
    except Exception:
        all_products = []
        
    if not all_products:
        return {
            "categories": [],
            "selected_category": category,
            "brands": [],
            "brands_dist": [],
            "capacities": [],
            "cells": [],
            "kpis": {"total_value": 0.0, "total_references": 0, "total_stock": 0, "coverage_pct": 0.0},
            "alerts": []
        }
        
    categories = sorted(list(set(p.get("category", "Otros") for p in all_products)))
    if not category and categories:
        category = categories[0]
        
    cat_products = [p for p in all_products if p.get("category") == category]
    
    # Aplicar filtro de color si se especifica
    if color_filter:
        cat_products = [p for p in cat_products if p.get("color", "") == color_filter]
    
    # Obtener colores disponibles en la categoría (antes de filtrar)
    all_cat_products = [p for p in all_products if p.get("category") == category]
    colors_available = sorted(list(set(p.get("color", "") for p in all_cat_products if p.get("color"))))
    
    # Determinar límites de precio para gamas (Económica, Media, Premium)
    pr_limits = []
    if category in ["Lavadoras", "Secadoras"]:
        pr_limits = [
            {"label": "Económica", "min": 0, "max": 350},
            {"label": "Media", "min": 350, "max": 550},
            {"label": "Premium", "min": 550, "max": 999999}
        ]
    elif category == "Lavavajillas":
        pr_limits = [
            {"label": "Económica", "min": 0, "max": 300},
            {"label": "Media", "min": 300, "max": 450},
            {"label": "Premium", "min": 450, "max": 999999}
        ]
    elif category == "Frigoríficos" or (category and ("frigo" in category.lower() or "frigorific" in category.lower())):
        pr_limits = [
            {"label": "Económica", "min": 0, "max": 400},
            {"label": "Media", "min": 400, "max": 700},
            {"label": "Premium", "min": 700, "max": 999999}
        ]
    else:
        pr_limits = [
            {"label": "Económica", "min": 0, "max": 200},
            {"label": "Media", "min": 200, "max": 400},
            {"label": "Premium", "min": 400, "max": 999999}
        ]
    
    # Obtener todas las marcas presentes en la categoría
    brands_in_cat = sorted(list(set(p.get("brand", "Genérico") for p in cat_products)))
    if "Genérico" in brands_in_cat:
        brands_in_cat.remove("Genérico")
        brands_in_cat.append("Genérico")
        
    capacities_set = set(p.get("capacity", "N/D") for p in cat_products)
    
    def sort_capacity(cap):
        if cap == "N/D":
            return (999, 0.0, 0.0, 0.0, "")
        
        # Si es dimensión de frigorífico de 2 componentes (ej: 177,5x56)
        dim_match_2 = re.match(r'^(\d+(?:[\.,]\d+)?)[Xx](\d+(?:[\.,]\d+)?)$', cap)
        if dim_match_2:
            try:
                h = float(dim_match_2.group(1).replace(",", "."))
                w = float(dim_match_2.group(2).replace(",", "."))
                return (0, h, w, 0.0, "")
            except ValueError:
                pass
        
        # Si es dimensión de frigorífico completa HxWxD (ej: 203X60X65)
        dim_match_3 = re.match(r'^(\d+(?:[\.,]\d+)?)[Xx](\d+(?:[\.,]\d+)?)[Xx](\d+(?:[\.,]\d+)?)$', cap)
        if dim_match_3:
            try:
                h = float(dim_match_3.group(1).replace(",", "."))
                w = float(dim_match_3.group(2).replace(",", "."))
                d = float(dim_match_3.group(3).replace(",", "."))
                return (0, h, w, d, "")
            except ValueError:
                pass
                
        # Fallback estándar para otros valores numéricos (ej: 8 kg)
        match = re.search(r'(\d+[\.,]?\d*)', cap)
        if match:
            try:
                val = float(match.group(1).replace(",", "."))
                return (1, val, 0.0, 0.0, "")
            except ValueError:
                pass
        return (100, 0.0, 0.0, 0.0, cap)
        
    capacities = sorted(list(capacities_set), key=sort_capacity)
    
    cells = []
    covered_cells_count = 0
    # Cada combinación marca-capacidad tiene 3 segmentos a cubrir (Eco, Media, Premium)
    total_segments_count = len(brands_in_cat) * len(capacities) * 3 if capacities else 0
    covered_segments_count = 0
    
    for cap in capacities:
        for brand_name in brands_in_cat:
            cell_products = [
                p for p in cat_products 
                if p.get("capacity") == cap and p.get("brand", "Genérico") == brand_name
            ]
            
            # Dividir en segmentos de precio
            eco_prods = [p for p in cell_products if pr_limits[0]["min"] <= p.get("cost", 0.0) < pr_limits[0]["max"]]
            med_prods = [p for p in cell_products if pr_limits[1]["min"] <= p.get("cost", 0.0) < pr_limits[1]["max"]]
            pre_prods = [p for p in cell_products if pr_limits[2]["min"] <= p.get("cost", 0.0) < pr_limits[2]["max"]]
            
            if len(eco_prods) > 0: covered_segments_count += 1
            if len(med_prods) > 0: covered_segments_count += 1
            if len(pre_prods) > 0: covered_segments_count += 1
            
            segments_data = {
                "E": {
                    "products": eco_prods,
                    "count": len(eco_prods),
                    "stock": sum(p.get("stock", 0) for p in eco_prods),
                    "status": "danger" if len(eco_prods) == 0 else ("warning" if len(eco_prods) == 1 else "success")
                },
                "M": {
                    "products": med_prods,
                    "count": len(med_prods),
                    "stock": sum(p.get("stock", 0) for p in med_prods),
                    "status": "danger" if len(med_prods) == 0 else ("warning" if len(med_prods) == 1 else "success")
                },
                "P": {
                    "products": pre_prods,
                    "count": len(pre_prods),
                    "stock": sum(p.get("stock", 0) for p in pre_prods),
                    "status": "danger" if len(pre_prods) == 0 else ("warning" if len(pre_prods) == 1 else "success")
                }
            }
            
            total_stock = sum(p.get("stock", 0) for p in cell_products)
            count = len(cell_products)
            status = "success" if count > 0 else "danger"
            
            cells.append({
                "capacity": cap,
                "brand": brand_name,
                "products": cell_products,
                "segments": segments_data,
                "count": count,
                "total_stock": total_stock,
                "status": status
            })
            
    total_value = sum(p.get("cost", 0.0) * p.get("stock", 0) for p in cat_products)
    total_references = len(cat_products)
    total_stock = sum(p.get("stock", 0) for p in cat_products)
    coverage_pct = (covered_segments_count / total_segments_count * 100) if total_segments_count > 0 else 0.0
    
    kpis = {
        "total_value": round(total_value, 2),
        "total_references": total_references,
        "total_stock": total_stock,
        "coverage_pct": round(coverage_pct, 1)
    }
    
    alerts = []
    for cap in capacities:
        for br in brands_in_cat:
            cell_products = [
                p for p in cat_products 
                if p.get("capacity") == cap and p.get("brand", "Genérico") == br
            ]
            eco_prods = [p for p in cell_products if pr_limits[0]["min"] <= p.get("cost", 0.0) < pr_limits[0]["max"]]
            med_prods = [p for p in cell_products if pr_limits[1]["min"] <= p.get("cost", 0.0) < pr_limits[1]["max"]]
            pre_prods = [p for p in cell_products if pr_limits[2]["min"] <= p.get("cost", 0.0) < pr_limits[2]["max"]]
            
            for label, prods in [("Económica", eco_prods), ("Media", med_prods), ("Premium", pre_prods)]:
                if not prods:
                    alerts.append({
                        "type": "danger",
                        "message": f"Falta gama: '{br}' en {cap} de gama {label}."
                    })
                elif sum(p.get("stock", 0) for p in prods) == 0:
                    alerts.append({
                        "type": "warning",
                        "message": f"Sin stock: '{br}' en {cap} de gama {label} sin unidades físicas."
                    })
                elif len(prods) == 1:
                    alerts.append({
                        "type": "info",
                        "message": f"Baja variedad: Solo tienes 1 opción de '{br}' en {cap} gama {label}."
                    })
                
    alerts.sort(key=lambda x: {"danger": 0, "warning": 1, "info": 2}[x["type"]])
    alerts = alerts[:10]
    
    # Calcular distribución por marcas para el panel lateral
    brands_dist = {}
    for p in cat_products:
        br = p.get("brand", "Genérico")
        stk = p.get("stock", 0)
        brands_dist[br] = brands_dist.get(br, 0) + stk
        
    brands_list = [{"brand": k, "stock": v} for k, v in brands_dist.items()]
    brands_list.sort(key=lambda x: x["stock"], reverse=True)
    
    return {
        "categories": categories,
        "selected_category": category,
        "selected_color": color_filter,
        "colors_available": colors_available,
        "brands": brands_in_cat,
        "brands_dist": brands_list,
        "capacities": capacities,
        "cells": cells,
        "kpis": kpis,
        "alerts": alerts
    }

@app.post("/api/stock/upload")
async def upload_stock_pdf(file: UploadFile = File(...)):
    stock_dir = os.path.join(DATA_DIR, "stock")
    os.makedirs(stock_dir, exist_ok=True)
    
    temp_pdf_path = os.path.join(stock_dir, "temp_inventory.pdf")
    inventory_file = os.path.join(stock_dir, "inventory.json")
    
    try:
        with open(temp_pdf_path, "wb") as f:
            content = await file.read()
            f.write(content)
            
        add_log("info", "PDF de stock subido. Iniciando extracción automática...")
        extracted_products = parse_erp_pdf(temp_pdf_path)
        
        if not extracted_products:
            raise HTTPException(status_code=400, detail="No se pudo extraer ningún producto del PDF. Comprueba el formato.")
            
        with open(inventory_file, "w", encoding="utf-8") as f:
            json.dump(extracted_products, f, indent=2, ensure_ascii=False)
            
        # Eliminar archivo temporal
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)
            
        add_log("success", f"Inventario ERP importado con éxito: {len(extracted_products)} referencias encontradas.")
        return {"status": "success", "count": len(extracted_products)}
    except Exception as e:
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)
        add_log("error", f"Error al procesar el PDF de inventario: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/stock/matrix")
async def get_stock_matrix(category: Optional[str] = None, color: Optional[str] = None):
    return get_stock_matrix_data(category, color_filter=color or "")


@app.get("/api/stock/export/xlsx")
async def export_stock_xlsx(category: Optional[str] = None, color: Optional[str] = None):
    """Genera y devuelve un informe Excel profesional de la categoría de stock seleccionada."""
    import io
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    
    data = get_stock_matrix_data(category or "Lavadoras", color_filter=color or "")
    
    wb = Workbook()
    ws = wb.active
    cat_label = data.get("selected_category", "Stock")
    ws.title = cat_label[:31]
    
    # ── Estilos ──────────────────────────────────────────────────────────────
    PURPLE      = "7C3AED"
    PURPLE_SOFT = "EDE9FE"
    DARK_ROW    = "1E1B4B"
    ALT_ROW     = "F5F3FF"
    WHITE       = "FFFFFF"
    GRAY_BORDER = "D1D5DB"
    SUCCESS_CLR = "059669"
    WARN_CLR    = "D97706"
    DANGER_CLR  = "DC2626"
    
    border_thin = Border(
        left=Side(style='thin', color=GRAY_BORDER),
        right=Side(style='thin', color=GRAY_BORDER),
        top=Side(style='thin', color=GRAY_BORDER),
        bottom=Side(style='thin', color=GRAY_BORDER)
    )
    
    def hdr_style(text, row, col, bold=True, bg=PURPLE, fg=WHITE, size=11):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=bold, color=fg, size=size, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        return cell
    
    def data_style(text, row, col, bold=False, bg=WHITE, fg="111827", align='left'):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=bold, color=fg, size=10, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        cell.border = border_thin
        return cell
    
    # ── Título del informe ───────────────────────────────────────────────────
    color_suffix = f" — Color: {color}" if color else ""
    title_text = f"Informe de Stock: {cat_label}{color_suffix}"
    date_text = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(bold=True, color=WHITE, size=16, name='Calibri')
    title_cell.fill = PatternFill("solid", fgColor=PURPLE)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('H1:J1')
    date_cell = ws['H1']
    date_cell.value = f"Generado: {date_text}"
    date_cell.font = Font(bold=False, color=WHITE, size=10, name='Calibri')
    date_cell.fill = PatternFill("solid", fgColor=PURPLE)
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # ── KPIs ─────────────────────────────────────────────────────────────────
    kpis = data.get("kpis", {})
    kpi_labels = ["Valoración de Stock", "Referencias Únicas", "Total Unidades", "Cobertura"]
    kpi_values = [
        f"{kpis.get('total_value', 0):,.2f} €",
        str(kpis.get('total_references', 0)),
        str(kpis.get('total_stock', 0)),
        f"{kpis.get('coverage_pct', 0)}%"
    ]
    ws.merge_cells('A2:J2')  # spacer
    ws['A2'].fill = PatternFill("solid", fgColor=PURPLE_SOFT)
    ws.row_dimensions[2].height = 10
    
    for i, (lbl, val) in enumerate(zip(kpi_labels, kpi_values)):
        col_start = i * 2 + 1
        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_start + 1)
        ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_start + 1)
        kpi_lbl_cell = ws.cell(row=3, column=col_start, value=lbl)
        kpi_lbl_cell.font = Font(bold=False, color="6B7280", size=9, name='Calibri')
        kpi_lbl_cell.fill = PatternFill("solid", fgColor=PURPLE_SOFT)
        kpi_lbl_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 18
        kpi_val_cell = ws.cell(row=4, column=col_start, value=val)
        kpi_val_cell.font = Font(bold=True, color=PURPLE, size=13, name='Calibri')
        kpi_val_cell.fill = PatternFill("solid", fgColor=PURPLE_SOFT)
        kpi_val_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 26
    
    # ── Hoja de la Matriz ────────────────────────────────────────────────────
    MATRIX_START_ROW = 7
    capacities = data.get("capacities", [])
    brands = data.get("brands", [])
    
    # Cabecera de la tabla
    hdr_style("Marca \\ Capacidad", MATRIX_START_ROW, 1, size=10)
    ws.column_dimensions['A'].width = 22
    for ci, cap in enumerate(capacities):
        col = ci + 2
        hdr_style(cap, MATRIX_START_ROW, col, size=9)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = max(10, len(cap) + 2)
    ws.row_dimensions[MATRIX_START_ROW].height = 30
    
    # Filas de marcas
    cells_lookup = {(c["brand"], c["capacity"]): c for c in data.get("cells", [])}
    for ri, brand in enumerate(brands):
        row = MATRIX_START_ROW + 1 + ri
        row_bg = ALT_ROW if ri % 2 == 0 else WHITE
        data_style(brand, row, 1, bold=True, bg=row_bg, fg="111827", align='left')
        ws.row_dimensions[row].height = 20
        for ci, cap in enumerate(capacities):
            col = ci + 2
            cell_data = cells_lookup.get((brand, cap))
            if cell_data and cell_data.get("count", 0) > 0:
                segs = cell_data.get("segments", {})
                e_ok = segs.get("E", {}).get("count", 0) > 0
                m_ok = segs.get("M", {}).get("count", 0) > 0
                p_ok = segs.get("P", {}).get("count", 0) > 0
                all_ok = e_ok and m_ok and p_ok
                none_ok = not e_ok and not m_ok and not p_ok
                total_stock = cell_data.get("total_stock", 0)
                count = cell_data.get("count", 0)
                cell_text = f"{count} ref / {total_stock} uds"
                bg = SUCCESS_CLR if all_ok else (WARN_CLR if not none_ok else DANGER_CLR)
                c = data_style(cell_text, row, col, bold=False, bg=row_bg, fg="111827", align='center')
            else:
                c = data_style("—", row, col, bg=row_bg, fg="9CA3AF", align='center')
    
    # ── Hoja de detalle de productos ─────────────────────────────────────────
    ws2 = wb.create_sheet(title="Detalle Productos")
    detail_headers = ["SKU/Modelo", "Marca", "Categoría", "Descripción", "Capacidad", "Color", "Stock", "Coste Unit. (€)", "Valor Total (€)"]
    for ci, h in enumerate(detail_headers):
        cell = ws2.cell(row=1, column=ci+1, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    ws2.row_dimensions[1].height = 28
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 8
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 14
    
    # Cargar datos del inventario completo para esta categoría
    inventory_file = os.path.join(DATA_DIR, "stock", "inventory.json")
    all_products = []
    if os.path.exists(inventory_file):
        try:
            with open(inventory_file, "r", encoding="utf-8") as f:
                all_products = json.load(f)
        except Exception:
            pass
    
    cat_prods = [p for p in all_products if p.get("category") == (category or "Lavadoras")]
    if color:
        cat_prods = [p for p in cat_prods if p.get("color", "") == color]
    cat_prods.sort(key=lambda p: (p.get("brand", ""), p.get("capacity", ""), p.get("sku", "")))
    
    for ri, p in enumerate(cat_prods):
        row = ri + 2
        row_bg = ALT_ROW if ri % 2 == 0 else WHITE
        cost = p.get("cost", 0.0)
        stk = p.get("stock", 0)
        vals = [p.get("sku", ""), p.get("brand", ""), p.get("category", ""), p.get("description", ""), p.get("capacity", ""), p.get("color", ""), stk, round(cost, 2), round(cost * stk, 2)]
        for ci, v in enumerate(vals):
            c = ws2.cell(row=row, column=ci+1, value=v)
            c.font = Font(size=9, name='Calibri', color="111827")
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal='center' if ci >= 6 else 'left', vertical='center')
            c.border = border_thin
        ws2.row_dimensions[row].height = 16
    
    # Auto-filter en hoja de detalle
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"
    
    # ── Generar respuesta de streaming ───────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_cat = re.sub(r'[^\w\-]', '_', cat_label)
    filename = f"informe_stock_{safe_cat}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/api/stock/raw")
async def get_stock_raw():
    inventory_file = os.path.join(DATA_DIR, "stock", "inventory.json")
    if not os.path.exists(inventory_file):
        return []
    try:
        with open(inventory_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

@app.post("/api/stock/clear")
async def clear_stock_data():
    inventory_file = os.path.join(DATA_DIR, "stock", "inventory.json")
    if os.path.exists(inventory_file):
        os.remove(inventory_file)
        add_log("info", "Datos de inventario eliminados correctamente.")
    return {"status": "success"}

if __name__ == "__main__":
    import uvicorn
    import os
    import sys
    port = int(os.environ.get("PORT", 8000))
    # En la versión online, enlazamos a 0.0.0.0 para ser accesibles externamente.
    # Desactivamos reload si se ejecuta compilado (PyInstaller) o en producción para evitar bucles.
    should_reload = not getattr(sys, 'frozen', False) and os.environ.get("ENV") != "production"
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=should_reload)
